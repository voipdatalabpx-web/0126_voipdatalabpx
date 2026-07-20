"""
IPLAN ASR Dashboard — Streamlit + Plotly
Fuente de datos: CSV locales (Maestros Entrante + Saliente)
Autor: Romina D'Angelo
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import os
from glob import glob
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Rutas locales (desarrollo en PC) ─────────────────────────────────────────
CARPETA_ENTRANTE = r"C:\Users\rdangelo\reportes_python_2026\darwin_ruta_entrante"
CARPETA_SALIENTE = r"C:\Users\rdangelo\reportes_python_2026\darwin_ruta_saliente"

# ── URLs públicas Google Sheets (producción Streamlit Cloud) ──────────────────
URL_ENTRANTE = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQhxNbRnZA5RKiKbbdUAwHNkhPGulFIIFYTFteGn3BwKy-qCIQo0dh7cgCDPepzYdUVm61-aYk8pLfo/pub?gid=0&single=true&output=csv"
URL_SALIENTE = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSv37aOQrLLuQEezWUr9oDpE5hrdtZvx1zCDLTLh7VANDpA4K1TqLFm4JsjI8Bw8qyPI7h9hVRPcffM/pub?gid=0&single=true&output=csv"
PROVEEDORES_CC   = ["266585-SAL", "DAINUS", "FRACAS", "ETROTTA", "AGINET", "NUETEL"]
LOGO_URL         = "https://www.iplan.com.ar/themes/contrib/iplan_b4/NEW/img/v2_LogoIplan.png"

C = {
    "navy":    "#001F5B", "primary": "#003087", "blue":    "#0055B3",
    "accent":  "#0077E6", "light":   "#3399FF", "cyan":    "#00C2FF",
    "success": "#00C853", "warning": "#FFB300", "danger":  "#FF1744",
    "bg":      "#050E24", "panel":   "#0A1A3A", "card":    "#0D2144",
    "border":  "#1A3560", "text":    "#E8F0FF", "muted":   "#7090C0",
    "white":   "#FFFFFF",
}
ASR_SCALE = [[0,"#FF1744"],[0.6,"#FFB300"],[0.85,"#00C853"],[1,"#00FF88"]]

# ── CSS ───────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Exo+2:wght@300;400;600;700;800&family=Rajdhani:wght@500;600;700&display=swap');

    /* Eliminar barra superior y padding de Streamlit */
    .stApp > header {{ display:none !important; }}
    #MainMenu, footer, header {{ visibility:hidden !important; display:none !important; }}
    .stDeployButton {{ display:none !important; }}
    .block-container {{
        padding-top: 2px !important;
        padding-bottom: 6px !important;
        max-width: 100% !important;
    }}
    .stMainBlockContainer {{ padding-top:2px !important; }}

    html, body, [class*="css"], .stApp {{
        font-family:'Exo 2',sans-serif;
        background-color:{C['bg']};
        color:{C['text']};
    }}

    /* ── Sidebar siempre visible 220px ── */
    [data-testid="stSidebar"],
    [data-testid="stSidebar"][aria-expanded="true"],
    [data-testid="stSidebar"][aria-expanded="false"] {{
        width:220px !important;
        min-width:220px !important;
        max-width:220px !important;
        transform:none !important;
        visibility:visible !important;
        display:block !important;
        background:linear-gradient(180deg,{C['navy']} 0%,{C['panel']} 100%);
        border-right:1px solid {C['border']};
    }}
    [data-testid="stSidebar"] > div,
    [data-testid="stSidebar"] > div:first-child {{
        width:220px !important;
        min-width:220px !important;
    }}
    /* Ocultar el botón de colapsar sidebar */
    [data-testid="collapsedControl"] {{ display:none !important; }}
    button[kind="header"] {{ display:none !important; }}
    [data-testid="stSidebar"] * {{ color:{C['text']} !important; }}
    [data-testid="stSidebar"] label {{
        color:{C['cyan']} !important; font-size:0.68rem; font-weight:700;
        text-transform:uppercase; letter-spacing:0.1em;
    }}

    /* Header compacto */
    .iplan-header {{
        background:linear-gradient(135deg,{C['navy']} 0%,{C['primary']} 50%,{C['blue']} 100%);
        border:1px solid {C['border']}; border-bottom:2px solid {C['cyan']};
        padding:5px 16px; border-radius:8px; margin-bottom:5px;
        display:flex; align-items:center; gap:12px; position:relative; overflow:hidden;
    }}
    .iplan-header::before {{
        content:''; position:absolute; top:-40px; right:-40px;
        width:160px; height:160px;
        background:radial-gradient(circle,rgba(0,194,255,.12) 0%,transparent 70%);
    }}
    .iplan-header img {{ height:24px; filter:brightness(0) invert(1); }}
    .iplan-header-text h1 {{
        color:{C['white']} !important; font-family:'Rajdhani',sans-serif;
        font-size:2.3rem; font-weight:700; margin:0;
        letter-spacing:0.06em; text-transform:uppercase; line-height:1;
    }}
    .iplan-header-text span {{ color:{C['cyan']}; font-size:0.75rem; letter-spacing:0.08em; }}
    .header-badge {{
        margin-left:auto; background:rgba(0,194,255,.15);
        border:1px solid {C['cyan']}; border-radius:20px; padding:2px 10px;
        font-size:0.65rem; color:{C['cyan']}; font-weight:700;
        letter-spacing:0.08em; text-transform:uppercase;
    }}

    /* KPI cards */
    .kpi-card {{
        background:{C['card']}; border:1px solid {C['border']};
        border-top:3px solid {C['accent']}; border-radius:10px;
        padding:10px 14px; margin-bottom:5px;
    }}
    .kpi-card.green  {{ border-top-color:{C['success']}; }}
    .kpi-card.orange {{ border-top-color:{C['warning']}; }}
    .kpi-card.red    {{ border-top-color:{C['danger']};  }}
    .kpi-card.cyan   {{ border-top-color:{C['cyan']};    }}
    .kpi-label {{
        font-size:0.6rem; font-weight:800; text-transform:uppercase;
        letter-spacing:0.14em; color:{C['cyan']}; margin-bottom:3px;
        text-shadow:0 0 8px rgba(0,194,255,0.4);
    }}
    .kpi-value {{
        font-family:'Rajdhani',sans-serif; font-size:2rem;
        font-weight:700; color:{C['white']}; line-height:1;
    }}
    .kpi-sub {{ font-size:0.65rem; color:{C['muted']}; margin-top:2px; font-weight:600; }}

    /* Section titles */
    .section-title {{
        font-family:'Rajdhani',sans-serif; font-size:1rem; font-weight:700;
        color:{C['cyan']}; text-transform:uppercase; letter-spacing:0.12em;
        border-bottom:1px solid {C['border']}; padding-bottom:3px; margin:7px 0 7px 0;
        display:flex; align-items:center; gap:8px;
    }}
    .section-title::before {{
        content:''; display:inline-block; width:3px; height:16px;
        background:{C['cyan']}; border-radius:2px; flex-shrink:0;
    }}

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {{
        background:{C['panel']}; border-radius:10px; padding:3px;
        border:1px solid {C['border']}; gap:3px; margin-top:4px;
    }}
    .stTabs [data-baseweb="tab"] {{
        font-family:'Rajdhani',sans-serif; font-weight:600; font-size:0.85rem;
        color:{C['muted']} !important; border-radius:7px; padding:5px 14px;
    }}
    .stTabs [aria-selected="true"] {{
        color:{C['cyan']} !important; background:{C['card']} !important;
        border:1px solid {C['border']};
    }}

    /* Stats sidebar */
    .stats-box {{
        background:rgba(0,119,230,.08); border:1px solid {C['border']};
        border-radius:8px; padding:8px; font-size:0.7rem;
        color:{C['muted']}; margin-top:10px; line-height:1.9;
    }}
    .stats-box b {{ color:{C['cyan']}; }}

    /* Button */
    .stButton button {{
        background:linear-gradient(135deg,{C['primary']},{C['accent']}) !important;
        color:white !important; border:none !important; border-radius:7px !important;
        font-family:'Rajdhani',sans-serif !important; font-weight:700 !important;
        font-size:0.82rem !important;
    }}

    /* Tabla fija con heatmap */
    .ht-wrap {{
        max-height:400px; overflow-y:auto; overflow-x:auto;
        border-radius:10px; border:1px solid {C['border']}; background:{C['panel']};
    }}
    .ht {{
        width:100%; border-collapse:collapse; font-size:0.75rem;
        font-family:'Exo 2',sans-serif; color:{C['text']};
    }}
    .ht thead tr th {{
        background:{C['primary']}; color:{C['cyan']} !important;
        font-family:'Rajdhani',sans-serif; font-weight:700; font-size:0.7rem;
        text-transform:uppercase; letter-spacing:0.08em;
        padding:7px 8px; text-align:left; position:sticky; top:0; z-index:2;
        border-bottom:2px solid {C['cyan']}; white-space:nowrap;
        cursor:pointer;
    }}
    .ht thead tr th:hover {{ background:{C['accent']}; }}
    .ht tbody tr {{ border-bottom:1px solid {C['border']}; }}
    .ht tbody tr:hover {{ background:rgba(0,119,230,0.1) !important; }}
    .ht tbody td {{ padding:5px 8px; white-space:nowrap; }}
    .ht tbody td.carrier-cell {{
        max-width:140px; overflow:hidden; text-overflow:ellipsis;
        font-weight:600; color:{C['light']};
    }}
    .ht tbody td.asr-cell {{ font-weight:700; color:{C['cyan']}; }}

    /* Insight box */
    .insight-box {{
        background:linear-gradient(135deg,rgba(0,48,135,0.4),rgba(0,194,255,0.08));
        border:1px solid {C['border']}; border-left:3px solid {C['cyan']};
        border-radius:8px; padding:12px 16px; margin:6px 0;
        font-size:0.8rem; color:{C['text']}; line-height:1.7;
    }}
    .insight-box b {{ color:{C['cyan']}; }}
    .insight-box .warn {{ color:{C['warning']}; font-weight:700; }}
    .insight-box .ok   {{ color:{C['success']}; font-weight:700; }}
    .insight-box .bad  {{ color:{C['danger']};  font-weight:700; }}
    </style>
    <script>
    function sortTable(tableId, col) {{
        var table = document.getElementById(tableId);
        var rows = Array.from(table.querySelectorAll('tbody tr'));
        var asc = table.getAttribute('data-sort-col') == col && table.getAttribute('data-sort-asc') == '1';
        rows.sort(function(a, b) {{
            var av = a.cells[col].innerText.replace('%','').replace(/[^0-9.,-]/g,'').replace(',','.');
            var bv = b.cells[col].innerText.replace('%','').replace(/[^0-9.,-]/g,'').replace(',','.');
            var an = parseFloat(av), bn = parseFloat(bv);
            if (!isNaN(an) && !isNaN(bn)) return asc ? an-bn : bn-an;
            return asc ? av.localeCompare(bv) : bv.localeCompare(av);
        }});
        var tbody = table.querySelector('tbody');
        rows.forEach(function(r){{ tbody.appendChild(r); }});
        table.setAttribute('data-sort-col', col);
        table.setAttribute('data-sort-asc', asc ? '0' : '1');
    }}
    </script>
    """, unsafe_allow_html=True)


# ── CARGA DE DATOS ─────────────────────────────────────────────────────────────
def _parsear_df(df, tipo, carrier_opts, ruta_opts):
    """Normaliza columnas, tipos y nombres de un DataFrame leído."""
    cols        = df.columns.tolist()
    col_carrier = next((c for c in carrier_opts if c in cols), None)
    col_ruta    = next((c for c in ruta_opts    if c in cols), None)
    if not col_carrier or not col_ruta:
        return None
    df = df.rename(columns={col_carrier:"CARRIER", col_ruta:"RUTA"})
    if "ASR_REAL" in df.columns and "ASR" not in df.columns:
        df = df.rename(columns={"ASR_REAL":"ASR"})
    df["TIPO"] = tipo
    df["ASR"]  = pd.to_numeric(
        df["ASR"].astype(str).str.replace("%","").str.strip(), errors="coerce")
    for col in ["CALLS_TOTAL","CALLS_OK","MINUTOS","AVG_CALL"]:
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0)
    df = df.dropna(subset=["CARRIER"])
    df = df[df["CARRIER"].astype(str).str.strip() != ""]
    return df[["FECHA","TIPO","CARRIER","RUTA","ASR",
               "CALLS_TOTAL","CALLS_OK","MINUTOS","AVG_CALL"]]


def _leer_csv(fuente, sep=";", decimal=","):
    """Lee un CSV desde ruta local o URL."""
    return pd.read_csv(fuente, encoding="utf-8-sig", sep=sep, decimal=decimal)


@st.cache_data(ttl=300, show_spinner=False)
def load_all_data(_unused=None):
    """
    Detecta el entorno automáticamente:
    - Streamlit Cloud: lee desde datos/ en el repo (GitHub)
    - PC local:        lee desde las carpetas Windows configuradas
    """
    import os as _os

    CARRIER_ENT = ["CARRIER_ORIGEN","CARRIER_DESTINO","CARRIER"]
    RUTA_ENT    = ["RUTA_ORIGEN","RUTA_DESTINO","RUTA"]
    CARRIER_SAL = ["CARRIER_DESTINO","CARRIER_ORIGEN","CARRIER"]
    RUTA_SAL    = ["RUTA_DESTINO","RUTA_ORIGEN","RUTA"]

    dfs = []

    # ── Ruta relativa al repo (Streamlit Cloud y también local si existe) ─────
    # Los CSV deben estar en la carpeta  datos/  del repositorio GitHub
    base_repo  = _os.path.dirname(_os.path.abspath(__file__))
    repo_ent   = _os.path.join(base_repo, "datos", "MAESTRO_ENTRANTE.csv")
    repo_sal   = _os.path.join(base_repo, "datos", "MAESTRO_SALIENTE.csv")

    fuentes_repo = [
        (repo_ent, "ENTRANTE", CARRIER_ENT, RUTA_ENT),
        (repo_sal, "SALIENTE", CARRIER_SAL, RUTA_SAL),
    ]

    repo_ok = True
    for ruta, tipo, c_opts, r_opts in fuentes_repo:
        if not _os.path.exists(ruta):
            repo_ok = False
            break
        try:
            df_raw = _leer_csv(ruta, sep=";", decimal=",")
            df = _parsear_df(df_raw, tipo, c_opts, r_opts)
            if df is not None and not df.empty:
                dfs.append(df)
            else:
                repo_ok = False
        except Exception as e:
            repo_ok = False
            st.warning(f"Error leyendo {tipo} desde repo: {e}")

    if dfs and repo_ok:
        return pd.concat(dfs, ignore_index=True)

    # ── Fallback: carpetas locales Windows ────────────────────────────────────
    dfs = []
    fuentes_local = [
        (CARPETA_ENTRANTE, "ENTRANTE", CARRIER_ENT, RUTA_ENT),
        (CARPETA_SALIENTE, "SALIENTE", CARRIER_SAL, RUTA_SAL),
    ]
    for carpeta, tipo, c_opts, r_opts in fuentes_local:
        if not _os.path.isdir(carpeta):
            continue
        sufijo   = "entrante" if tipo == "ENTRANTE" else "saliente"
        latest   = _os.path.join(carpeta, f"maestro_{sufijo}_LATEST.csv")
        archivos = [latest] if _os.path.exists(latest) else glob(_os.path.join(carpeta, "*.csv"))
        for archivo in archivos:
            try:
                df_raw = _leer_csv(archivo, sep=";", decimal=",")
                df = _parsear_df(df_raw, tipo, c_opts, r_opts)
                if df is not None:
                    dfs.append(df)
            except Exception as e:
                st.warning(f"Error leyendo {_os.path.basename(archivo)}: {e}")

    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


# ── HELPERS ───────────────────────────────────────────────────────────────────
def fmt_num(n,dec=0):
    if pd.isna(n): return "-"
    return f"{n:,.{dec}f}".replace(",","X").replace(".",",").replace("X",".")

def fmt_asr(v):
    if pd.isna(v): return "-"
    return f"{int(round(v))}%"

def fmt_compact(n):
    if pd.isna(n): return "-"
    n = float(n)
    if abs(n)>=1_000_000: return f"{n/1_000_000:.1f}M"
    if abs(n)>=1_000:     return f"{n/1_000:.1f}K"
    return f"{int(round(n))}"

def kpi_card(label,value,sub="",color="blue"):
    cls = {"green":"green","orange":"orange","red":"red","cyan":"cyan"}.get(color,"")
    st.markdown(f"""
    <div class="kpi-card {cls}">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
    </div>""", unsafe_allow_html=True)

def section(title):
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)

def base(fig,title="",height=360):
    fig.update_layout(
        title=dict(text=title,font=dict(family="Rajdhani",size=13,color=C["cyan"])),
        paper_bgcolor=C["card"], plot_bgcolor=C["panel"],
        font=dict(family="Exo 2",size=11,color=C["text"]),
        margin=dict(l=8,r=8,t=32,b=8), height=height,
        legend=dict(bgcolor="rgba(0,0,0,0)",borderwidth=0,
                    font=dict(color=C["text"]),
                    orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
    )
    fig.update_xaxes(showgrid=True,gridcolor=C["border"],
                     linecolor=C["border"],tickfont=dict(color=C["muted"],size=11))
    fig.update_yaxes(showgrid=True,gridcolor=C["border"],
                     linecolor=C["border"],tickfont=dict(color=C["muted"],size=11))
    return fig

def bar_h(df_plot,x_col,y_col,label_col,title,height=400,y_color=None):
    """Barra horizontal con etiquetas en blanco resaltado en eje Y."""
    df_s = df_plot.sort_values(x_col,ascending=True)
    fig = go.Figure(go.Bar(
        x=df_s[x_col], y=df_s[y_col], orientation="h",
        text=df_s[label_col], textposition="outside",
        textfont=dict(size=9,color=C["text"]),
        marker=dict(color=df_s["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,
                    showscale=False),
    ))
    fig.update_layout(yaxis=dict(
        tickfont=dict(size=9,color=C["white"],family="Exo 2"),
    ))
    return base(fig,title,height)

def tabla_fija(df_t, table_id, calls_col="CALLS_TOTAL", asr_col="ASR",
               carrier_col="CARRIER", titulo=""):
    """
    Tabla HTML fija con scroll, heatmap en CALLS, ordenable por click en cabecera,
    carrier truncado, ASR en cyan, titulos en negrita.
    """
    if titulo:
        st.markdown(f'<div class="section-title">{titulo}</div>', unsafe_allow_html=True)

    cols = df_t.columns.tolist()
    # Calcular max para heatmap sobre valor numérico original
    calls_idx = cols.index(calls_col) if calls_col in cols else -1

    # Construir cabeceras con onclick para ordenar
    thead_cells = ""
    for i,c in enumerate(cols):
        thead_cells += f'<th onclick="sortTable(\'{table_id}\',{i})" title="Click para ordenar">{c}</th>'

    # Calcular max_calls a partir del texto (quitar separadores)
    def parse_num(s):
        try: return float(str(s).replace(".","").replace(",","."))
        except: return 0.0

    max_c = 0
    if calls_col in cols:
        max_c = max(parse_num(v) for v in df_t[calls_col])

    def heat_bg(val_str):
        v = parse_num(val_str)
        if max_c==0: return "transparent"
        ratio = min(v/max_c,1.0)
        return f"rgba(0,119,230,{ratio*0.55:.2f})"

    rows_html = ""
    for _,row in df_t.iterrows():
        cells = ""
        for c in cols:
            val = row[c]
            if c == calls_col:
                bg = heat_bg(val)
                cells += f'<td style="background:{bg};font-weight:700;">{val}</td>'
            elif c == asr_col:
                cells += f'<td class="asr-cell">{val}</td>'
            elif c == carrier_col:
                short = str(val)[:22]+"…" if len(str(val))>22 else str(val)
                cells += f'<td class="carrier-cell" title="{val}">{short}</td>'
            else:
                cells += f"<td>{val}</td>"
        rows_html += f"<tr>{cells}</tr>"

    st.markdown(f"""
    <div class="ht-wrap">
    <table class="ht" id="{table_id}">
        <thead><tr>{thead_cells}</tr></thead>
        <tbody>{rows_html}</tbody>
    </table>
    </div>""", unsafe_allow_html=True)

def gen_insights(filtro):
    """Genera insights automáticos basados en los datos."""
    ent = filtro[filtro["TIPO"]=="ENTRANTE"]
    sal = filtro[filtro["TIPO"]=="SALIENTE"]
    insights = []

    # ASR global
    asr_g = filtro["ASR"].mean()
    if asr_g < 70:
        insights.append(f'<span class="bad">ASR global critico: {fmt_asr(asr_g)}</span> — revisar rutas con mayor trafico.')
    elif asr_g < 85:
        insights.append(f'<span class="warn">ASR global por debajo del target: {fmt_asr(asr_g)}</span> — target minimo 85%.')
    else:
        insights.append(f'<span class="ok">ASR global dentro del target: {fmt_asr(asr_g)}</span>')

    # Rutas criticas entrantes
    if not ent.empty:
        rutas_malas_ent = (ent.groupby("RUTA").agg(ASR=("ASR","mean"),CALLS=("CALLS_TOTAL","sum"))
                           .query("ASR < 70 and CALLS > 1000").sort_values("CALLS",ascending=False))
        if not rutas_malas_ent.empty:
            lista = ", ".join([f"<b>{r}</b> ({fmt_asr(a)})" for r,a in
                               zip(rutas_malas_ent.index[:3], rutas_malas_ent["ASR"].iloc[:3])])
            insights.append(f'Rutas ENT con ASR &lt;70% y alto trafico: {lista}')

    # Rutas criticas salientes
    if not sal.empty:
        rutas_malas_sal = (sal.groupby("RUTA").agg(ASR=("ASR","mean"),CALLS=("CALLS_TOTAL","sum"))
                           .query("ASR < 70 and CALLS > 1000").sort_values("CALLS",ascending=False))
        if not rutas_malas_sal.empty:
            lista = ", ".join([f"<b>{r}</b> ({fmt_asr(a)})" for r,a in
                               zip(rutas_malas_sal.index[:3], rutas_malas_sal["ASR"].iloc[:3])])
            insights.append(f'Rutas SAL con ASR &lt;70% y alto trafico: {lista}')

    # Top carrier por volumen
    if not ent.empty:
        top = ent.groupby("CARRIER")["MINUTOS"].sum().nlargest(1)
        insights.append(f'Mayor volumen ENT: <b>{top.index[0]}</b> con {fmt_compact(top.iloc[0])} minutos.')

    if not sal.empty:
        top_sal = sal.groupby("CARRIER")["MINUTOS"].sum().nlargest(1)
        insights.append(f'Mayor volumen SAL: <b>{top_sal.index[0]}</b> con {fmt_compact(top_sal.iloc[0])} minutos.')

    return insights


# ── DIARIO — Paleta y helpers estilo "Darwin Daily Report" (Excel) ────────────
DIARIO_C = {
    "navy_title":  "#1F3864", "navy_sub":  "#0D1B2A", "navy_sub_txt": "#AAAAAA",
    "green_title": "#375623", "green_sub": "#1A3A0F",
    "orange_title":"#C55A11", "orange_sub":"#7B2D00",
    "blue_header": "#2F5597",
    "row_alt": "#EBF3FB", "row_base": "#FFFFFF", "row_txt": "#1A1A2E",
    "ok_bg": "#D9F2D9", "warn_bg": "#FFF2CC", "bad_bg": "#FFD7D7",
    "ok_txt": "#1A6B1A", "warn_txt": "#8A6D00", "bad_txt": "#B00020",
}

def estado_asr(v):
    """Etiqueta + colores de semaforo segun ASR, igual criterio que el reporte Excel Darwin."""
    if pd.isna(v):
        return "-", "#F0F0F0", DIARIO_C["row_txt"]
    if v >= 80:
        return "✅ BUENO", DIARIO_C["ok_bg"], DIARIO_C["ok_txt"]
    if v >= 50:
        return "⚠️ REGULAR", DIARIO_C["warn_bg"], DIARIO_C["warn_txt"]
    return "🔴 BAJO", DIARIO_C["bad_bg"], DIARIO_C["bad_txt"]

def _diario_title(texto, sub, bg, sub_bg, sub_txt="#AAAAAA"):
    """Barra de titulo + subtitulo estilo Excel (fondo de color, texto centrado)."""
    st.markdown(f"""
    <div style="background:{bg};border-radius:8px 8px 0 0;padding:10px 16px;text-align:center;">
        <span style="color:#FFFFFF;font-family:'Rajdhani',sans-serif;font-size:1.3rem;
                     font-weight:800;letter-spacing:.04em;text-transform:uppercase;">{texto}</span>
    </div>
    <div style="background:{sub_bg};border-radius:0 0 8px 8px;padding:5px 16px;
                text-align:center;margin-bottom:12px;">
        <span style="color:{sub_txt};font-size:0.78rem;">{sub}</span>
    </div>
    """, unsafe_allow_html=True)

def _diario_tabla(df_t, header_bg, cols_map, id_tabla, estado_col=True):
    """
    Tabla HTML estilo Excel Darwin: cabecera de color solido, filas alternadas
    blanco/celeste, columna ESTADO con semaforo (BUENO/REGULAR/BAJO).
    cols_map: lista de tuplas (col_df, etiqueta_visible, formatter_o_None)
    """
    thead = "".join(
        f'<th style="background:{header_bg} !important;color:#FFFFFF;">{lbl}</th>'
        for _, lbl, _ in cols_map
    )
    if estado_col:
        thead += f'<th style="background:{header_bg} !important;color:#FFFFFF;">ESTADO</th>'

    rows = ""
    for i, (_, row) in enumerate(df_t.iterrows()):
        bg = DIARIO_C["row_alt"] if i % 2 == 0 else DIARIO_C["row_base"]
        cells = ""
        for col, _, fmt in cols_map:
            val = fmt(row[col]) if fmt else row[col]
            cells += f'<td style="background:{bg};color:{DIARIO_C["row_txt"]};">{val}</td>'
        if estado_col:
            txt, ebg, etxt = estado_asr(row["ASR"])
            cells += (f'<td style="background:{ebg};color:{etxt};font-weight:700;'
                      f'text-align:center;">{txt}</td>')
        rows += f"<tr>{cells}</tr>"

    st.markdown(f"""
    <div class="ht-wrap" style="max-height:480px;">
    <table class="ht" id="{id_tabla}">
        <thead><tr>{thead}</tr></thead>
        <tbody>{rows}</tbody>
    </table></div>
    """, unsafe_allow_html=True)


# ── PÁGINAS ───────────────────────────────────────────────────────────────────


def fmt_dia_semana(fecha_series):
    """Formatea fechas como 'Lun 02', 'Mar 03', etc."""
    dias = {0:"Lun",1:"Mar",2:"Mie",3:"Jue",4:"Vie",5:"Sab",6:"Dom"}
    return [f"{dias[d.weekday()]} {d.strftime('%d')}" for d in pd.to_datetime(fecha_series)]

# Paleta fija para todos los gráficos de evolución
_COLOR_CALLS   = "#1E90FF"          # azul dodger  — Llamadas
_COLOR_CALLS_F = "rgba(30,144,255,0.13)"   # sombra celeste
_COLOR_MIN     = "#FF2040"          # rojo vibrante — Minutos
_COLOR_MIN_F   = "rgba(255,32,64,0.10)"    # sombra rojo
_COLOR_ASR     = "#00E676"          # verde neón    — ASR
_COLOR_ASR_F   = "rgba(0,230,118,0.10)"    # sombra verde

def grafico_evo_combo(d_evo, color_bar=None, color_minutos=None, color_asr=None,
                      titulo="", height=400, solo_lineas=False, mini=False):
    """
    Gráfico combinado: Llamadas (azul) + Minutos (rojo vibrante) + ASR (verde).
    Todos con área rellena transparente debajo de la línea.
    solo_lineas=True: Llamadas como línea (no barra).
    mini=True: fuentes reducidas para cuadrícula.
    Eje X: día semana abreviado (Lun 02, Mar 03…).
    """
    # Colores fijos independiente de lo que se pase
    c_calls = _COLOR_CALLS
    c_min   = _COLOR_MIN
    c_asr   = _COLOR_ASR

    x_labels = fmt_dia_semana(d_evo["FECHA"])
    sz_txt  = 11 if mini else 14     # triple del anterior (~5→14, ~7→11 mini)
    sz_mrk  = 6  if mini else 9
    sz_line = 2  if mini else 3

    fig = go.Figure()

    if not solo_lineas:
        # Barras Llamadas — borde gris plata grueso
        fig.add_trace(go.Bar(
            x=x_labels, y=d_evo["CALLS"], name="Llamadas",
            marker=dict(
                color=c_calls, opacity=0.25,
                line=dict(color="#B0BEC5", width=2.5),
            ),
            text=d_evo["CALLS"].map(fmt_compact), textposition="outside",
            textfont=dict(size=sz_txt, color="#FFFFFF", family="Rajdhani"),
            yaxis="y1",
        ))
    else:
        # Línea Llamadas — azul con sombra celeste
        fig.add_trace(go.Scatter(
            x=x_labels, y=d_evo["CALLS"], name="Llamadas",
            mode="lines+markers+text",
            line=dict(color=c_calls, width=sz_line+1),
            marker=dict(size=sz_mrk+1, color=c_calls,
                        line=dict(color="#FFFFFF", width=1.5)),
            yaxis="y1",
            fill="tozeroy", fillcolor=_COLOR_CALLS_F,
            text=d_evo["CALLS"].map(fmt_compact), textposition="top center",
            textfont=dict(size=sz_txt, color="#FFFFFF", family="Rajdhani"),
        ))

    # Línea Minutos — rojo vibrante + área roja transparente
    fig.add_trace(go.Scatter(
        x=x_labels, y=d_evo["MINUTOS"], name="Minutos",
        mode="lines+markers+text",
        line=dict(color=c_min, width=sz_line+1),
        marker=dict(size=sz_mrk+1, color=c_min,
                    line=dict(color="#FFFFFF", width=1.5)),
        yaxis="y1",
        fill="tozeroy", fillcolor=_COLOR_MIN_F,
        text=d_evo["MINUTOS"].map(fmt_compact), textposition="top center",
        textfont=dict(size=sz_txt, color="#FFFFFF", family="Rajdhani"),
    ))

    # Línea ASR — verde neón + área verde transparente + punteada
    fig.add_trace(go.Scatter(
        x=x_labels, y=d_evo["ASR"], name="ASR %",
        mode="lines+markers+text",
        line=dict(color=c_asr, width=sz_line, dash="dot"),
        marker=dict(size=sz_mrk, color=c_asr, symbol="diamond",
                    line=dict(color="#FFFFFF", width=1.5)),
        yaxis="y2",
        fill="tozeroy", fillcolor=_COLOR_ASR_F,
        text=d_evo["ASR"].map(fmt_asr), textposition="top center",
        textfont=dict(size=sz_txt, color="#FFFFFF", family="Rajdhani"),
    ))

    fig.update_layout(
        yaxis=dict(
            title="" if mini else "Llamadas / Minutos",
            showgrid=True, gridcolor="rgba(26,53,96,0.5)",
            tickfont=dict(color="#7090C0", size=9 if mini else 11),
            zeroline=False,
        ),
        yaxis2=dict(
            title="" if mini else "ASR %",
            overlaying="y", side="right",
            range=[0, 118], ticksuffix="%",
            tickfont=dict(color=c_asr, size=9 if mini else 11),
            showgrid=False,
        ),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.02,
            xanchor="right", x=1,
            font=dict(color="#E8F0FF", size=10 if mini else 13, family="Rajdhani"),
            bgcolor="rgba(0,0,0,0)",
        ),
        barmode="overlay",
        xaxis=dict(
            tickfont=dict(color="#E8F0FF", size=10 if mini else 12, family="Rajdhani"),
            showgrid=False, linecolor="#1A3560",
        ),
        margin=dict(l=4 if mini else 8, r=4 if mini else 8,
                    t=32 if mini else 44, b=4 if mini else 8),
    )
    fig.add_hline(
        y=85, line_dash="dot", line_color=c_asr, line_width=1.5,
        annotation_text="85%" if mini else "Target ASR 85%",
        annotation_font=dict(color=c_asr, size=9 if mini else 11),
        yref="y2",
    )
    return fig

def page_resumen(df, filtro):
    ent = filtro[filtro["TIPO"]=="ENTRANTE"]
    sal = filtro[filtro["TIPO"]=="SALIENTE"]

    section("KPIs — Indicadores Clave del Mes  (Acumulado del periodo seleccionado)")
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1: kpi_card("Calls ENT", fmt_compact(ent["CALLS_TOTAL"].sum()),
                       "Acum: " + fmt_num(ent["CALLS_TOTAL"].sum()))
    with c2: kpi_card("Calls SAL", fmt_compact(sal["CALLS_TOTAL"].sum()),
                       "Acum: " + fmt_num(sal["CALLS_TOTAL"].sum()),"cyan")
    with c3: kpi_card("Minutos ENT", fmt_compact(ent["MINUTOS"].sum()),
                       "Acum: " + fmt_num(ent["MINUTOS"].sum(),0))
    with c4: kpi_card("Minutos SAL", fmt_compact(sal["MINUTOS"].sum()),
                       "Acum: " + fmt_num(sal["MINUTOS"].sum(),0),"cyan")
    asr_g = filtro["ASR"].mean()
    with c5: kpi_card("ASR Global", fmt_asr(asr_g), "promedio todas las rutas",
                       "green" if asr_g>=85 else "orange" if asr_g>=60 else "red")
    with c6: kpi_card("Rutas Activas",
                       f"ENT: {ent['RUTA'].nunique()}  SAL: {sal['RUTA'].nunique()}",
                       f"Dias analizados: {filtro['FECHA'].nunique()}","cyan")

    section("KPIs — Volumen Minutos por Cliente — Top 5 ENT  (Acumulado del periodo)")
    top5 = (ent.groupby("CARRIER")
            .agg(MINUTOS=("MINUTOS","sum"),CALLS=("CALLS_TOTAL","sum"),ASR=("ASR","mean"))
            .nlargest(5,"MINUTOS").reset_index())
    cols5 = st.columns(5)
    clrs5 = ["blue","cyan","blue","cyan","blue"]
    for i,(_,row) in enumerate(top5.iterrows()):
        nom = row["CARRIER"][:22]+"…" if len(row["CARRIER"])>22 else row["CARRIER"]
        with cols5[i]:
            kpi_card(nom,fmt_compact(row["MINUTOS"]),
                     f"ASR {fmt_asr(row['ASR'])} · {fmt_compact(row['CALLS'])} calls",clrs5[i])

    # Evolucion Diaria ENT — Llamadas · Minutos · ASR (combo)
    section("Evolucion Diaria ENT — Llamadas · Minutos · ASR")
    d_evo_ent = (ent.groupby("FECHA")
                 .agg(CALLS=("CALLS_TOTAL","sum"), MINUTOS=("MINUTOS","sum"), ASR=("ASR","mean"))
                 .reset_index().sort_values("FECHA"))
    if not d_evo_ent.empty:
        fig_evo_ent = grafico_evo_combo(
            d_evo_ent,
            color_bar="#0077E6",
            color_minutos="#FF3B4E",
            color_asr="#4DA6FF",
            titulo="",
            height=420,
        )
        st.plotly_chart(
            base(fig_evo_ent, "Evolucion Diaria ENT — Llamadas · Minutos (rojo) · ASR % (verde punteado)", 420),
            use_container_width=True,
        )

    # Evolucion Minutos
    section("Evolucion Diaria — Minutos")
    col1,col2 = st.columns(2)
    with col1:
        d3 = (filtro.groupby(["FECHA","TIPO"]).agg(MINUTOS=("MINUTOS","sum"))
              .reset_index().sort_values("FECHA"))
        d3["TL"] = d3["TIPO"].map({"ENTRANTE":"ENT","SALIENTE":"SAL"})
        d3["lbl"] = d3["MINUTOS"].map(fmt_compact)
        fig3 = px.bar(d3,x="FECHA",y="MINUTOS",color="TL",barmode="group",text="lbl",
                      color_discrete_map={"ENT":C["warning"],"SAL":C["success"]},
                      labels={"MINUTOS":"Minutos","FECHA":"","TL":"Tipo"})
        fig3.update_traces(textposition="outside",textfont=dict(size=9,color=C["text"]))
        st.plotly_chart(base(fig3,"Minutos por Dia — ENT & SAL"),use_container_width=True)
    with col2:
        d2 = (filtro.groupby(["FECHA","TIPO"]).agg(CALLS=("CALLS_TOTAL","sum"))
              .reset_index().sort_values("FECHA"))
        d2["TL"] = d2["TIPO"].map({"ENTRANTE":"ENT","SALIENTE":"SAL"})
        d2["lbl"] = d2["CALLS"].map(fmt_compact)
        fig2 = go.Figure()
        for tk,ck,fc in [("ENT",C["warning"],"rgba(255,179,0,0.15)"),
                          ("SAL",C["success"],"rgba(0,200,83,0.12)")]:
            sub = d2[d2["TL"]==tk].sort_values("FECHA")
            fig2.add_trace(go.Scatter(
                x=sub["FECHA"],y=sub["CALLS"],mode="lines+markers+text",
                name=tk,text=sub["lbl"],textposition="top center",
                textfont=dict(size=9,color=C["text"]),
                line=dict(color=ck,width=2),marker=dict(size=6,color=ck),
                fill="tozeroy",fillcolor=fc))
        st.plotly_chart(base(fig2,"Llamadas por Dia — ENT & SAL"),use_container_width=True)

    # Evolucion ASR
    section("Evolucion Diaria — ASR")
    col3,col4 = st.columns(2)
    with col3:
        d = (filtro.groupby(["FECHA","TIPO"]).agg(ASR=("ASR","mean"))
             .reset_index().sort_values("FECHA"))
        d["TL"] = d["TIPO"].map({"ENTRANTE":"ENT","SALIENTE":"SAL"})
        d["lbl"] = d["ASR"].map(fmt_asr)
        fig = px.line(d,x="FECHA",y="ASR",color="TL",markers=True,text="lbl",
                      color_discrete_map={"ENT":C["warning"],"SAL":C["success"]},
                      labels={"ASR":"ASR","FECHA":"","TL":"Tipo"})
        fig.update_traces(textposition="top center",textfont=dict(size=10,color=C["text"]))
        fig.update_yaxes(ticksuffix="%",range=[0,110])
        fig.add_hline(y=85,line_dash="dot",line_color=C["success"],
                      annotation_text="Target 85%",
                      annotation_font=dict(color=C["success"],size=10))
        st.plotly_chart(base(fig,"ASR por Dia — ENT & SAL"),use_container_width=True)
    with col4:
        d4 = (filtro.groupby(["FECHA","TIPO"]).agg(ASR=("ASR","mean"))
              .reset_index().sort_values("FECHA"))
        d4["TL"] = d4["TIPO"].map({"ENTRANTE":"ENT","SALIENTE":"SAL"})
        d4["lbl"] = d4["ASR"].map(fmt_asr)
        fig4 = px.bar(d4,x="FECHA",y="ASR",color="TL",barmode="group",text="lbl",
                      color_discrete_map={"ENT":C["warning"],"SAL":C["success"]},
                      labels={"ASR":"ASR","FECHA":"","TL":"Tipo"})
        fig4.update_traces(textposition="outside",textfont=dict(size=9,color=C["text"]))
        fig4.update_yaxes(ticksuffix="%",range=[0,115])
        fig4.add_hline(y=85,line_dash="dot",line_color=C["success"])
        st.plotly_chart(base(fig4,"ASR Comparativo ENT vs SAL"),use_container_width=True)

    # Top 10
    section("Top 10 Rutas — Llamadas")
    col5,col6 = st.columns(2)
    with col5:
        te = (ent.groupby("RUTA").agg(CALLS=("CALLS_TOTAL","sum"),ASR=("ASR","mean"))
              .nlargest(10,"CALLS").reset_index())
        te["lbl"] = te["CALLS"].map(fmt_compact)
        st.plotly_chart(bar_h(te,"CALLS","RUTA","lbl","Top 10 ENT — Llamadas",400),
                        use_container_width=True)
    with col6:
        ts = (sal.groupby("RUTA").agg(CALLS=("CALLS_TOTAL","sum"),ASR=("ASR","mean"))
              .nlargest(10,"CALLS").reset_index())
        ts["lbl"] = ts["CALLS"].map(fmt_compact)
        st.plotly_chart(bar_h(ts,"CALLS","RUTA","lbl","Top 10 SAL — Llamadas",400),
                        use_container_width=True)

    # Insights
    section("Insights Detectados")
    insights = gen_insights(filtro)
    for ins in insights:
        st.markdown(f'<div class="insight-box">{ins}</div>', unsafe_allow_html=True)


def page_clientes(df, filtro):
    # ── datos del periodo filtrado ─────────────────────────────────────────────
    ent = filtro[filtro["TIPO"]=="ENTRANTE"]
    if ent.empty:
        st.info("Sin datos de rutas entrantes."); return

    # Determinar mes actual y mes anterior desde el filtro
    fechas_unicas = pd.to_datetime(ent["FECHA"]).dt.to_period("M").unique()
    mes_actual    = fechas_unicas.max()
    mes_anterior  = mes_actual - 1

    ent_act = ent[pd.to_datetime(ent["FECHA"]).dt.to_period("M") == mes_actual]
    ent_ant = df[(df["TIPO"]=="ENTRANTE") &
                 (pd.to_datetime(df["FECHA"]).dt.to_period("M") == mes_anterior)]

    # ── KPIs Clientes ──────────────────────────────────────────────────────────
    section("KPIs — Trafico Entrante (Acumulado del periodo)")
    c1,c2,c3,c4 = st.columns(4)
    asr_ent = ent["ASR"].mean()
    with c1: kpi_card("Calls ENT", fmt_compact(ent["CALLS_TOTAL"].sum()),
                       "Acum: " + fmt_num(ent["CALLS_TOTAL"].sum()))
    with c2: kpi_card("Minutos ENT", fmt_compact(ent["MINUTOS"].sum()),
                       "Acum: " + fmt_num(ent["MINUTOS"].sum(),0), "cyan")
    with c3: kpi_card("ASR ENT", fmt_asr(asr_ent), "promedio acumulado",
                       "green" if asr_ent>=85 else "orange" if asr_ent>=60 else "red")
    with c4: kpi_card("Clientes Activos", str(ent["CARRIER"].nunique()),
                       f"Rutas: {ent['RUTA'].nunique()}", "cyan")

    # ── Evolucion Llamadas + Minutos + ASR ────────────────────────────────────
    section("Evolucion Diaria ENT — Llamadas · Minutos · ASR")
    d_evo = (ent.groupby("FECHA")
              .agg(CALLS=("CALLS_TOTAL","sum"), MINUTOS=("MINUTOS","sum"), ASR=("ASR","mean"))
              .reset_index().sort_values("FECHA"))
    fig_evo = grafico_evo_combo(
        d_evo,
        color_bar="#0077E6",    # azul eléctrico — borde gris plata
        color_minutos="#FF3B4E",
        color_asr="#4DA6FF",
        titulo="",
        height=420,
    )
    st.plotly_chart(base(fig_evo, "Evolucion Diaria ENT — Llamadas · Minutos (rojo) · ASR % (azul punteado)", 420),
                    use_container_width=True)

    # ── Top Clientes: Mes Actual vs Mes Anterior (LLAMADAS, orden mayor a menor) ──
    top_act = (ent_act.groupby("CARRIER")
               .agg(CALLS_ACT=("CALLS_TOTAL","sum"), MIN_ACT=("MINUTOS","sum"),
                    ASR_ACT=("ASR","mean")).reset_index())
    top_ant = (ent_ant.groupby("CARRIER")
               .agg(CALLS_ANT=("CALLS_TOTAL","sum"), MIN_ANT=("MINUTOS","sum"),
                    ASR_ANT=("ASR","mean")).reset_index())

    cmp = top_act.merge(top_ant, on="CARRIER", how="outer").fillna(0)
    cmp = cmp.sort_values("CALLS_ACT", ascending=False).head(10)
    cmp["delta_calls"] = cmp["CALLS_ACT"] - cmp["CALLS_ANT"]
    cmp["delta_pct"]   = cmp.apply(
        lambda r: f"+{fmt_compact(r['delta_calls'])}" if r["delta_calls"]>=0
                  else fmt_compact(r["delta_calls"]), axis=1)

    # KPIs delta encima del gráfico
    section(f"Top Clientes — {mes_actual} vs {mes_anterior} — Llamadas Acumuladas (mayor a menor)")
    cols_delta = st.columns(min(5, len(cmp)))
    for i, (_, row) in enumerate(cmp.head(5).iterrows()):
        nom = row["CARRIER"][:20]+"…" if len(row["CARRIER"])>20 else row["CARRIER"]
        color_d = "green" if row["delta_calls"]>=0 else "red"
        with cols_delta[i]:
            kpi_card(
                nom,
                fmt_compact(row["CALLS_ACT"]),
                f"Llamadas acum · vs ant: {row['delta_pct']} · ASR {fmt_asr(row['ASR_ACT'])}",
                color_d,
            )

    # Gráfico comparativo — orden mayor a menor (sort asc para barras h)
    cmp_plot = cmp.sort_values("CALLS_ACT", ascending=True)
    fig_cmp = go.Figure()
    fig_cmp.add_trace(go.Bar(
        y=cmp_plot["CARRIER"], x=cmp_plot["CALLS_ACT"], name=f"Llamadas {mes_actual}",
        orientation="h",
        marker=dict(color="#FF6B00", opacity=0.92,
                    line=dict(color="#FFB347", width=1.5)),
        text=cmp_plot["CALLS_ACT"].map(fmt_compact), textposition="outside",
        textfont=dict(size=10, color="#FFFFFF", family="Rajdhani"),
    ))
    fig_cmp.add_trace(go.Bar(
        y=cmp_plot["CARRIER"], x=cmp_plot["CALLS_ANT"], name=f"Llamadas {mes_anterior}",
        orientation="h",
        marker=dict(color="#78909C", opacity=0.65,
                    line=dict(color="#B0BEC5", width=1)),
        text=cmp_plot["CALLS_ANT"].map(fmt_compact), textposition="outside",
        textfont=dict(size=9, color="#CFD8DC", family="Rajdhani"),
    ))
    fig_cmp.update_layout(
        barmode="group",
        yaxis=dict(tickfont=dict(size=9, color="#FFFFFF", family="Exo 2")),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1, font=dict(color="#E8F0FF", family="Rajdhani")),
    )
    st.plotly_chart(
        base(fig_cmp, f"Llamadas por Cliente — {mes_actual} vs {mes_anterior}", 440),
        use_container_width=True,
    )

    # ── Gráficos originales ────────────────────────────────────────────────────
    section("Distribucion por Cliente")
    pc = (ent.groupby("CARRIER")
          .agg(CALLS_TOTAL=("CALLS_TOTAL","sum"),CALLS_OK=("CALLS_OK","sum"),
               MINUTOS=("MINUTOS","sum"),ASR=("ASR","mean"),RUTAS=("RUTA","nunique"))
          .reset_index().sort_values("CALLS_TOTAL",ascending=False))
    pc["lbl"] = pc["CALLS_TOTAL"].map(fmt_compact)

    col1,col2 = st.columns([3,2])
    with col1:
        fig_c = go.Figure(go.Bar(
            x=pc.head(20).sort_values("CALLS_TOTAL")["CALLS_TOTAL"],
            y=pc.head(20).sort_values("CALLS_TOTAL")["CARRIER"],
            orientation="h",
            text=pc.head(20).sort_values("CALLS_TOTAL")["lbl"],
            textposition="outside",
            textfont=dict(size=9,color=C["text"]),
            marker=dict(color=pc.head(20).sort_values("CALLS_TOTAL")["ASR"],
                        colorscale=ASR_SCALE,cmin=0,cmax=100,showscale=False),
        ))
        fig_c.update_layout(yaxis=dict(tickfont=dict(size=9,color="#FFFFFF",family="Exo 2")))
        st.plotly_chart(base(fig_c,"Llamadas por Carrier Entrante",460),use_container_width=True)
    with col2:
        fig2 = px.pie(pc.head(10),values="MINUTOS",names="CARRIER",
                      color_discrete_sequence=[C["accent"],C["cyan"],C["light"],C["blue"],
                                               C["warning"],C["success"],"#6B5CE7",
                                               "#FF6B6B","#4ECDC4","#45B7D1"])
        fig2.update_traces(
            textposition="inside",textinfo="percent+label",
            textfont=dict(size=11,color=C["white"],family="Rajdhani"),
            insidetextfont=dict(size=11,color=C["white"],family="Rajdhani"),
            hole=0.35,pull=[0.03]*min(10,len(pc)))
        fig2.update_layout(showlegend=False)
        st.plotly_chart(base(fig2,"Distribucion Minutos",460),use_container_width=True)

    # Tabla
    carrier_sel = st.selectbox("Filtrar por Carrier",
                                ["Todos"]+sorted(ent["CARRIER"].unique().tolist()))
    df_det = ent if carrier_sel=="Todos" else ent[ent["CARRIER"]==carrier_sel]
    tabla = (df_det.groupby(["CARRIER","RUTA"])
             .agg(ASR=("ASR","mean"),CALLS_TOTAL=("CALLS_TOTAL","sum"),
                  CALLS_OK=("CALLS_OK","sum"),MINUTOS=("MINUTOS","sum"),
                  AVG_CALL=("AVG_CALL","mean"))
             .reset_index().sort_values("CALLS_TOTAL",ascending=False))
    tabla["ASR"]         = tabla["ASR"].map(fmt_asr)
    tabla["CALLS_TOTAL"] = tabla["CALLS_TOTAL"].map(fmt_num)
    tabla["CALLS_OK"]    = tabla["CALLS_OK"].map(fmt_num)
    tabla["MINUTOS"]     = tabla["MINUTOS"].map(lambda x: fmt_num(x,2))
    tabla["AVG_CALL"]    = tabla["AVG_CALL"].map(lambda x: fmt_num(x,2))
    tabla_fija(tabla,"tbl_clientes",titulo="Detalle por Ruta Entrante")


def page_salientes(df, filtro):
    sal = filtro[filtro["TIPO"]=="SALIENTE"]
    if sal.empty:
        st.info("Sin datos de rutas salientes."); return

    # ── KPIs Proveedores ───────────────────────────────────────────────────────
    section("KPIs — Trafico Saliente / Proveedores (Acumulado del periodo)")
    c1,c2,c3,c4 = st.columns(4)
    asr_sal = sal["ASR"].mean()
    with c1: kpi_card("Calls SAL", fmt_compact(sal["CALLS_TOTAL"].sum()),
                       "Acum: " + fmt_num(sal["CALLS_TOTAL"].sum()), "cyan")
    with c2: kpi_card("Minutos SAL", fmt_compact(sal["MINUTOS"].sum()),
                       "Acum: " + fmt_num(sal["MINUTOS"].sum(),0))
    with c3: kpi_card("ASR SAL", fmt_asr(asr_sal), "promedio acumulado",
                       "green" if asr_sal>=85 else "orange" if asr_sal>=60 else "red")
    with c4: kpi_card("Proveedores Activos", str(sal["CARRIER"].nunique()),
                       f"Rutas: {sal['RUTA'].nunique()}", "cyan")

    # ── Evolucion Llamadas + Minutos + ASR punteado ────────────────────────────
    section("Evolucion Diaria — Llamadas, Minutos y ASR (SAL)")
    d_evo = (sal.groupby("FECHA")
              .agg(CALLS=("CALLS_TOTAL","sum"), MINUTOS=("MINUTOS","sum"), ASR=("ASR","mean"))
              .reset_index().sort_values("FECHA"))

    fig_evo = grafico_evo_combo(
        d_evo,
        color_bar="#00C853",    # verde para línea Llamadas SAL
        color_minutos="#FF3B4E",
        color_asr="#4DA6FF",
        titulo="",
        height=420,
        solo_lineas=True,       # sin barras, solo líneas
    )
    st.plotly_chart(base(fig_evo, "Evolucion Diaria SAL — Llamadas (verde) · Minutos (rojo) · ASR % (azul punteado)", 420),
                    use_container_width=True)

    # ── Graficos de distribución ───────────────────────────────────────────────
    section("Distribucion por Proveedor")
    pc = (sal.groupby("CARRIER")
          .agg(CALLS_TOTAL=("CALLS_TOTAL","sum"),CALLS_OK=("CALLS_OK","sum"),
               MINUTOS=("MINUTOS","sum"),ASR=("ASR","mean"))
          .reset_index().sort_values("CALLS_TOTAL",ascending=False))
    pc["lbl"] = pc["CALLS_TOTAL"].map(fmt_compact)

    col1,col2 = st.columns([3,2])
    with col1:
        pc_s = pc.head(20).sort_values("CALLS_TOTAL")
        fig_s = go.Figure(go.Bar(
            x=pc_s["CALLS_TOTAL"], y=pc_s["CARRIER"], orientation="h",
            text=pc_s["lbl"], textposition="outside",
            textfont=dict(size=9,color=C["text"]),
            marker=dict(color=pc_s["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,
                        showscale=False),
        ))
        fig_s.update_layout(yaxis=dict(tickfont=dict(size=9,color="#FFFFFF",family="Exo 2")))
        st.plotly_chart(base(fig_s,"Llamadas por Proveedor Saliente",460),use_container_width=True)
    with col2:
        pc_t = pc.head(15).copy()
        pc_t["lbl_tm"] = pc_t["MINUTOS"].map(fmt_compact)
        fig2 = go.Figure(go.Treemap(
            labels=pc_t["CARRIER"], parents=[""] * len(pc_t),
            values=pc_t["MINUTOS"], text=pc_t["lbl_tm"],
            texttemplate="<b>%{label}</b><br>%{text}",
            textfont=dict(size=12,color=C["white"],family="Rajdhani"),
            marker=dict(colors=pc_t["ASR"],
                        colorscale=[[0,"#FF1744"],[0.6,"#FFB300"],[0.85,"#00C853"],[1,"#00FF88"]],
                        cmin=0, cmax=100, showscale=False),
        ))
        st.plotly_chart(base(fig2,"Minutos por Proveedor (Treemap)",460),use_container_width=True)

    # ── Resumen detallado por carrier/ruta — todos, ordenado mayor a menor ──────
    section("Resumen Detallado de Desempeno por Carrier — Todas las Rutas")

    reporte = (sal.groupby(["CARRIER","RUTA"])
               .agg(
                   ASR=("ASR","mean"),
                   CALLS_TOTAL=("CALLS_TOTAL","sum"),
                   CALLS_OK=("CALLS_OK","sum"),
                   MINUTOS=("MINUTOS","sum"),
                   AVG_CALL=("AVG_CALL","mean"),
               )
               .reset_index()
               .sort_values("CALLS_TOTAL", ascending=False)
               .reset_index(drop=True))

    reporte.index = reporte.index + 1

    def asr_bullet(v):
        if pd.isna(v): return "-"
        color = "#00C853" if v>=85 else ("#FFB300" if v>=60 else "#FF1744")
        return f'<span style="color:{color};font-weight:700;">{int(round(v))}%</span>'

    def calls_bar_html(calls, max_calls):
        if max_calls == 0: return fmt_num(calls)
        pct = min(int(calls / max_calls * 100), 100)
        return (
            '<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="width:60px;background:#1A3560;border-radius:3px;height:8px;">'
            f'<div style="width:{pct}%;background:#1E90FF;border-radius:3px;height:8px;"></div>'
            '</div>'
            f'<span>{fmt_num(calls)}</span>'
            '</div>'
        )

    max_calls = reporte["CALLS_TOTAL"].max()
    cols_tabla = ["N°","CARRIER","RUTA","ASR REAL (%)","TOTAL CALLS","CALLS OK","MINUTOS","AVG CALL"]
    thead = ""
    for i, c in enumerate(cols_tabla):
        thead += f'<th onclick="sortTable(\'tbl_reporte_sal\',{i})" title="Ordenar">{c}</th>'

    rows_html = ""
    for idx, row in reporte.iterrows():
        carrier = str(row["CARRIER"])
        nom = carrier[:28]+"\u2026" if len(carrier)>28 else carrier
        asr_v   = row["ASR"]
        # dots de color ASR (3 circulos: verde si >=85, naranja si >=60, rojo si <60)
        def dot(threshold):
            if pd.isna(asr_v): return '<span style="color:#444;">●</span>'
            color = "#00C853" if asr_v>=85 else ("#FFB300" if asr_v>=60 else "#FF1744")
            return f'<span style="color:{color};">●</span>'
        dots = dot(85) + dot(70) + dot(50)
        asr_html = f'{dots} <b style="color:#FFFFFF;">{int(round(asr_v)) if not pd.isna(asr_v) else "-"}%</b>'
        rows_html += (
            "<tr>"
            f'<td style="color:#7090C0;text-align:center;font-weight:700;">{idx}</td>'
            f'<td style="color:#FFFFFF;font-weight:700;max-width:160px;overflow:hidden;text-overflow:ellipsis;" title="{carrier}">{nom}</td>'
            f'<td style="color:#A0C0E0;">{row["RUTA"]}</td>'
            f'<td style="text-align:center;">{asr_html}</td>'
            f'<td>{calls_bar_html(row["CALLS_TOTAL"], max_calls)}</td>'
            f'<td style="color:#B0C8E8;">{fmt_num(row["CALLS_OK"])}</td>'
            f'<td style="color:#FF3B4E;font-weight:700;">{fmt_num(row["MINUTOS"],2)}</td>'
            f'<td style="color:#7090C0;">{fmt_num(row["AVG_CALL"],2)}</td>'
            "</tr>"
        )

    st.markdown(
        '<div class="ht-wrap" style="max-height:560px;">'
        '<table class="ht" id="tbl_reporte_sal">'
        f'<thead><tr>{thead}</tr></thead>'
        f'<tbody>{rows_html}</tbody>'
        '</table></div>',
        unsafe_allow_html=True,
    )


def page_poi(df, filtro):
    section("POIs — IPLAN")
    # Solo SALIENTES para el heatmap y análisis
    poi_all = filtro[filtro["RUTA"].str.contains("POI",case=False,na=False)]
    poi_sal = poi_all[poi_all["TIPO"]=="SALIENTE"]

    if poi_all.empty:
        st.info("No se encontraron rutas POI."); return

    # Solo SALIENTES en los gráficos
    pp = (poi_sal.groupby("RUTA")
          .agg(ASR=("ASR","mean"),CALLS_TOTAL=("CALLS_TOTAL","sum"),
               CALLS_OK=("CALLS_OK","sum"),MINUTOS=("MINUTOS","sum"))
          .reset_index().sort_values("CALLS_TOTAL",ascending=False))

    if pp.empty:
        st.info("Sin datos Salientes para rutas POI."); return

    col1,col2 = st.columns(2)
    with col1:
        dc = pp.sort_values("CALLS_TOTAL")
        dc["lbl"] = dc["CALLS_TOTAL"].map(fmt_compact)
        fig = go.Figure(go.Bar(
            x=dc["CALLS_TOTAL"], y=dc["RUTA"], orientation="h",
            text=dc["lbl"], textposition="outside",
            textfont=dict(size=9,color=C["text"]),
            marker=dict(color=dc["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,showscale=False),
        ))
        fig.update_layout(yaxis=dict(tickfont=dict(size=9,color=C["white"])))
        st.plotly_chart(base(fig,"Llamadas POI — Salientes",430),use_container_width=True)
    with col2:
        da = pp.sort_values("ASR")
        da["lbl"] = da["ASR"].map(fmt_asr)
        fig2 = go.Figure(go.Bar(
            x=da["ASR"], y=da["RUTA"], orientation="h",
            text=da["lbl"], textposition="outside",
            textfont=dict(size=9,color=C["text"]),
            marker=dict(color=da["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,showscale=False),
        ))
        fig2.update_xaxes(ticksuffix="%",range=[0,115])
        fig2.update_layout(yaxis=dict(tickfont=dict(size=9,color=C["white"])))
        fig2.add_vline(x=85,line_dash="dot",line_color=C["success"])
        st.plotly_chart(base(fig2,"ASR POI — Salientes",430),use_container_width=True)

    # Heatmap solo SALIENTES
    if not poi_sal.empty and filtro["FECHA"].nunique()>1:
        section("Heatmap ASR POI por Fecha — Solo Salientes")
        pivot = poi_sal.pivot_table(index="RUTA",columns="FECHA",values="ASR",aggfunc="mean")
        fig3 = px.imshow(pivot,color_continuous_scale=["#FF1744","#FFB300","#00C853"],
                         range_color=[0,100],aspect="auto",text_auto=".0f",
                         labels=dict(color="ASR%"))
        fig3.update_traces(textfont=dict(size=9,color="white"))
        st.plotly_chart(base(fig3,"Heatmap ASR POIs Salientes",400),use_container_width=True)

    # Tabla fija POIs
    tabla = pp.copy()
    tabla["ASR"]         = tabla["ASR"].map(fmt_asr)
    tabla["CALLS_TOTAL"] = tabla["CALLS_TOTAL"].map(fmt_num)
    tabla["MINUTOS"]     = tabla["MINUTOS"].map(lambda x: fmt_num(x,2))
    tabla_fija(tabla,"tbl_poi",carrier_col="RUTA",titulo="Detalle POIs")

    # Insights POIs
    section("Insights POIs Detectados")
    if not poi_sal.empty:
        malos = (poi_sal.groupby("RUTA").agg(ASR=("ASR","mean"),CALLS=("CALLS_TOTAL","sum"))
                 .query("ASR < 80").sort_values("ASR"))
        if not malos.empty:
            for ruta,row in malos.iterrows():
                nivel = "bad" if row["ASR"]<70 else "warn"
                st.markdown(f'<div class="insight-box">POI SAL <b>{ruta}</b>: '
                            f'<span class="{nivel}">ASR {fmt_asr(row["ASR"])}</span> '
                            f'con {fmt_compact(row["CALLS"])} llamadas.</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown('<div class="insight-box"><span class="ok">Todos los POIs Salientes con ASR aceptable.</span></div>',
                        unsafe_allow_html=True)


def page_prefix_cc(df, filtro):
    section("Prefix Call Center — Proveedores NUETEL")
    st.caption("Rutas: 266585-SAL · DAINUS · FRACAS · ETROTTA · AGINET · NUETEL")

    mask = (filtro["RUTA"].str.upper().str.contains("|".join(PROVEEDORES_CC),na=False) |
            filtro["CARRIER"].str.upper().str.contains("|".join(PROVEEDORES_CC),na=False))
    cc = filtro[mask]
    if cc.empty:
        st.info("Sin datos para los proveedores CC."); return

    pr = (cc.groupby(["CARRIER","RUTA"])
          .agg(ASR=("ASR","mean"),CALLS_TOTAL=("CALLS_TOTAL","sum"),
               CALLS_OK=("CALLS_OK","sum"),MINUTOS=("MINUTOS","sum"),
               AVG_CALL=("AVG_CALL","mean"))
          .reset_index().sort_values("CALLS_TOTAL",ascending=False))

    # KPIs: Calls · Minutos · ASR — Acumulado del mes
    asr_cc = cc["ASR"].mean()
    c1,c2,c3 = st.columns(3)
    with c1: kpi_card("Calls Totales CC",fmt_compact(cc["CALLS_TOTAL"].sum()),
                       "Acum: " + fmt_num(cc["CALLS_TOTAL"].sum()))
    with c2: kpi_card("Minutos CC",fmt_compact(cc["MINUTOS"].sum()),
                       "Acum: " + fmt_num(cc["MINUTOS"].sum(),0),"cyan")
    with c3: kpi_card("ASR CC",fmt_asr(asr_cc),"Promedio acumulado del mes",
                       "green" if asr_cc>=85 else "orange" if asr_cc>=60 else "red")

    # KPIs por proveedor (Minutos acumulados)
    section("Minutos por Proveedor — Acumulado del Mes")
    por_prov = (cc.groupby("CARRIER")
                .agg(MINUTOS=("MINUTOS","sum"),CALLS=("CALLS_TOTAL","sum"),ASR=("ASR","mean"))
                .reset_index().sort_values("MINUTOS",ascending=False))
    cols_prov = st.columns(min(len(por_prov),6))
    for i,(_,row) in enumerate(por_prov.head(6).iterrows()):
        nom = row["CARRIER"][:20]+"…" if len(row["CARRIER"])>20 else row["CARRIER"]
        with cols_prov[i]:
            kpi_card(nom, fmt_compact(row["MINUTOS"]),
                     f"Acum · ASR {fmt_asr(row['ASR'])} · {fmt_compact(row['CALLS'])} calls",
                     "cyan" if i%2==0 else "blue")

    # ── Evolución Minutos + Llamadas por día por proveedor — 2 columnas ────────
    section("Evolucion Minutos Mensual por Proveedor — Minutos (rojo) · Llamadas por Dia")
    proveedores_lista = por_prov["CARRIER"].tolist()
    n_prov  = len(proveedores_lista)
    n_cols  = 2                                      # 2 por fila
    n_rows  = (n_prov + n_cols - 1) // n_cols
    # Colores para la línea de LLAMADAS de cada proveedor (minutos siempre rojo)
    colores_calls = [C["cyan"], C["warning"], C["success"], C["light"],
                     "#A78BFA", "#34D399", "#FB923C", C["danger"]]

    for fila in range(n_rows):
        grid_cols = st.columns(n_cols)
        for col_idx in range(n_cols):
            prov_idx = fila * n_cols + col_idx
            if prov_idx >= n_prov:
                break
            prov       = proveedores_lista[prov_idx]
            color_call = colores_calls[prov_idx % len(colores_calls)]

            df_prov = (cc[cc["CARRIER"]==prov]
                       .groupby("FECHA").agg(
                           MINUTOS=("MINUTOS","sum"),
                           CALLS=("CALLS_TOTAL","sum"),
                           ASR=("ASR","mean"))
                       .reset_index().sort_values("FECHA"))

            nom_short = prov[:32]+"…" if len(prov)>32 else prov

            # Mini gráfico: Llamadas como línea de color, Minutos en rojo, ASR azul
            fig_p = grafico_evo_combo(
                df_prov,
                color_bar=color_call,
                color_minutos="#FF3B4E",
                color_asr="#4DA6FF",
                titulo=nom_short,
                height=300,
                solo_lineas=True,
                mini=True,
            )
            with grid_cols[col_idx]:
                st.plotly_chart(
                    base(fig_p, nom_short, height=300),
                    use_container_width=True
                )

    # ── Concentración de Proveedores top por volumen de Minutos (MES) ─────────
    section("Concentracion de Proveedores top por volumen de Minutos (MES)")
    col1,col2 = st.columns(2)
    with col1:
        pr_c = pr.sort_values("CALLS_TOTAL",ascending=True)
        pr_c["lbl"] = pr_c["CALLS_TOTAL"].map(fmt_compact)
        fig = go.Figure(go.Bar(
            x=pr_c["CALLS_TOTAL"], y=pr_c["RUTA"], orientation="h",
            text=pr_c["lbl"], textposition="outside",
            textfont=dict(size=13,color="#FFFFFF",family="Rajdhani"),
            marker=dict(color=pr_c["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,
                        showscale=False),
        ))
        fig.update_layout(yaxis=dict(tickfont=dict(size=13,color="#FFFFFF",family="Rajdhani")))
        st.plotly_chart(base(fig,"Llamadas por Ruta — CC"),use_container_width=True)
    with col2:
        da = pr.sort_values("ASR",ascending=True)
        da["lbl"] = da["ASR"].map(fmt_asr)
        fig2 = go.Figure(go.Bar(
            x=da["ASR"], y=da["RUTA"], orientation="h",
            text=da["lbl"], textposition="outside",
            textfont=dict(size=13,color="#FFFFFF",family="Rajdhani"),
            marker=dict(color=da["ASR"],colorscale=ASR_SCALE,cmin=0,cmax=100,
                        showscale=False),
        ))
        fig2.update_layout(yaxis=dict(tickfont=dict(size=13,color="#FFFFFF",family="Rajdhani")))
        fig2.update_xaxes(ticksuffix="%",range=[0,115])
        fig2.add_vline(x=85,line_dash="dot",line_color=C["success"],
                       annotation_text="Target 85%",
                       annotation_font=dict(color=C["success"],size=11))
        st.plotly_chart(base(fig2,"ASR por Ruta — CC"),use_container_width=True)

    # ── Resumen detallado CC — todos los proveedores, mayor a menor ─────────────
    section("Resumen Detallado de Desempeno por Proveedor — Prefix CC")

    reporte_cc = (cc.groupby(["CARRIER","RUTA"])
                  .agg(
                      ASR=("ASR","mean"),
                      CALLS_TOTAL=("CALLS_TOTAL","sum"),
                      CALLS_OK=("CALLS_OK","sum"),
                      MINUTOS=("MINUTOS","sum"),
                      AVG_CALL=("AVG_CALL","mean"),
                  )
                  .reset_index()
                  .sort_values("CALLS_TOTAL", ascending=False)
                  .reset_index(drop=True))

    reporte_cc.index = reporte_cc.index + 1

    def asr_dots_cc(v):
        if pd.isna(v): return "-"
        color = "#00C853" if v>=85 else ("#FFB300" if v>=60 else "#FF1744")
        filled  = int(v / 100 * 5)
        dots_on  = f'<span style="color:{color};">{"●" * filled}</span>'
        dots_off = f'<span style="color:#2A3A5A;">{"●" * (5-filled)}</span>'
        return f'{dots_on}{dots_off} <b style="color:#FFFFFF;">{int(round(v))}%</b>'

    def bar_calls_cc(calls, max_c):
        if max_c == 0: return fmt_num(calls)
        pct = min(int(calls / max_c * 100), 100)
        return (
            '<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="width:55px;background:#1A3560;border-radius:3px;height:8px;">'
            f'<div style="width:{pct}%;background:#1E90FF;border-radius:3px;height:8px;"></div>'
            '</div>'
            f'<span style="color:#FFFFFF;font-weight:600;">{fmt_num(calls)}</span>'
            '</div>'
        )

    max_c = reporte_cc["CALLS_TOTAL"].max()
    cols_h = ["N°","CARRIER","RUTA","ASR REAL (%)","TOTAL CALLS","CALLS OK","MINUTOS","AVG CALL"]

    thead_cc = ""
    for i,c in enumerate(cols_h):
        thead_cc += f'<th onclick="sortTable(\'tbl_cc_reporte\',{i})" title="Ordenar">{c}</th>'

    rows_cc = ""
    for idx, row in reporte_cc.iterrows():
        carrier = str(row["CARRIER"])
        nom = carrier[:26]+"\u2026" if len(carrier)>26 else carrier
        rows_cc += (
            "<tr>"
            f'<td style="color:#7090C0;text-align:center;font-weight:700;">{idx}</td>'
            f'<td style="color:#FFFFFF;font-weight:700;max-width:150px;overflow:hidden;text-overflow:ellipsis;" title="{carrier}">{nom}</td>'
            f'<td style="color:#A0C0E0;">{row["RUTA"]}</td>'
            f'<td style="text-align:left;white-space:nowrap;">{asr_dots_cc(row["ASR"])}</td>'
            f'<td>{bar_calls_cc(row["CALLS_TOTAL"], max_c)}</td>'
            f'<td style="color:#B0C8E8;">{fmt_num(row["CALLS_OK"])}</td>'
            f'<td style="color:#FF3B4E;font-weight:700;">{fmt_num(row["MINUTOS"],2)}</td>'
            f'<td style="color:#7090C0;">{fmt_num(row["AVG_CALL"],2)}</td>'
            "</tr>"
        )

    st.markdown(
        '<div class="ht-wrap" style="max-height:520px;">'
        '<table class="ht" id="tbl_cc_reporte">'
        f'<thead><tr>{thead_cc}</tr></thead>'
        f'<tbody>{rows_cc}</tbody>'
        '</table></div>',
        unsafe_allow_html=True,
    )

    # ── Insights al final de la página ───────────────────────────────────────
    section("Insights — Proveedores CC")
    for _,row in por_prov.iterrows():
        prov  = row["CARRIER"]
        asr_p = row["ASR"]
        min_p = row["MINUTOS"]
        cal_p = row["CALLS"]
        nivel = "ok" if asr_p>=85 else ("warn" if asr_p>=70 else "bad")
        st.markdown(
            f'<div class="insight-box"><b>{prov}</b> — '
            f'Minutos acum: <b>{fmt_compact(min_p)}</b> · '
            f'Llamadas: <b>{fmt_compact(cal_p)}</b> · '
            f'ASR: <span class="{nivel}">{fmt_asr(asr_p)}</span></div>',
            unsafe_allow_html=True,
        )


def _diario_insights(sal, ent):
    """Genera las secciones de Insights & Alertas replicando el reporte Excel Darwin."""
    def box_section(titulo):
        st.markdown(
            f'<div style="border-left:4px solid {DIARIO_C["navy_title"]};'
            f'padding:6px 12px;font-weight:800;color:{C["cyan"]};margin:16px 0 8px;'
            f'text-transform:uppercase;font-family:Rajdhani,sans-serif;font-size:0.9rem;">'
            f'{titulo}</div>', unsafe_allow_html=True)

    def item(titulo, texto, bg=None):
        style = f'background:{bg};' if bg else ""
        st.markdown(
            f'<div class="insight-box" style="{style}"><b>{titulo}</b><br>'
            f'<span style="font-size:0.82rem;">{texto}</span></div>',
            unsafe_allow_html=True)

    calls_sal, ok_sal, min_sal = sal["CALLS_TOTAL"].sum(), sal["CALLS_OK"].sum(), sal["MINUTOS"].sum()
    calls_ent, ok_ent, min_ent = ent["CALLS_TOTAL"].sum(), ent["CALLS_OK"].sum(), ent["MINUTOS"].sum()
    asr_sal = sal["ASR"].mean() if not sal.empty else float("nan")
    asr_ent = ent["ASR"].mean() if not ent.empty else float("nan")

    box_section("📋 Resumen Ejecutivo del Día")
    item("📞 Volumen total de tráfico",
         f"Se procesaron {fmt_num(calls_sal)} intentos de llamadas salientes, completando "
         f"{fmt_num(ok_sal)} llamadas exitosas ({fmt_num(min_sal,0)} minutos). "
         f"El ASR promedio saliente fue {fmt_asr(asr_sal)} y entrante {fmt_asr(asr_ent)}.")
    item("📡 Cobertura de rutas",
         f"Se registraron {sal['RUTA'].nunique()} rutas salientes activas y "
         f"{ent['RUTA'].nunique()} rutas entrantes. El tráfico está distribuido entre "
         f"múltiples carriers.")

    box_section("🚨 Alertas Críticas · ASR Bajo")
    hubo_alerta = False
    for tipo, data in [("Saliente", sal), ("Entrante", ent)]:
        if data.empty:
            continue
        malas = (data.groupby("RUTA").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                 .query("ASR < 30 and CALLS > 100").sort_values("CALLS", ascending=False))
        if not malas.empty:
            hubo_alerta = True
            lista = " | ".join(f"{r} ({fmt_asr(a)})" for r, a in
                                zip(malas.index[:5], malas["ASR"].iloc[:5]))
            item(f"⚡ {tipo}: {len(malas)} rutas con ASR &lt; 30%",
                 f"Rutas de alto volumen con ASR crítico: {lista}", bg="rgba(255,23,68,0.10)")
    if not hubo_alerta:
        item("✅ Sin alertas críticas de ASR", "No se detectaron rutas de alto volumen con ASR &lt; 30%.")

    box_section("🏆 Rutas Destacadas · Mejor Performance")
    hubo_destacada = False
    for tipo, data in [("SAL", sal), ("ENT", ent)]:
        if data.empty:
            continue
        buenas = (data.groupby(["RUTA", "CARRIER"])
                  .agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"), MIN=("MINUTOS", "sum"))
                  .query("ASR >= 90 and CALLS > 500")
                  .sort_values("CALLS", ascending=False).head(3))
        for (ruta, carrier), row in buenas.iterrows():
            hubo_destacada = True
            item(f"✅ {tipo} · {ruta}",
                 f"Carrier: {carrier} | ASR: {fmt_asr(row['ASR'])} | {fmt_num(row['CALLS'])} llamadas | "
                 f"{fmt_num(row['MIN'],0)} min — Ruta estable y confiable.", bg="rgba(0,200,83,0.10)")
    if not hubo_destacada:
        item("Sin rutas destacadas", "Ninguna ruta supera el umbral de ASR ≥90% con volumen &gt;500 llamadas hoy.")

    box_section("📊 Análisis por Carrier Principal")
    if not sal.empty:
        top_c = (sal.groupby("CARRIER")
                 .agg(CALLS=("CALLS_TOTAL", "sum"), MIN=("MINUTOS", "sum"), ASR=("ASR", "mean"))
                 .nlargest(3, "MIN"))
        for carrier, row in top_c.iterrows():
            item(f"📶 {carrier} · Saliente",
                 f"{fmt_num(row['CALLS'])} llamadas | {fmt_num(row['MIN'],0)} min | "
                 f"ASR promedio: {fmt_asr(row['ASR'])}")

    box_section("💡 Recomendaciones Operativas")
    recs = []
    if not sal.empty:
        peor = (sal.groupby("CARRIER").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                .query("CALLS > 1000").nsmallest(1, "ASR"))
        if not peor.empty:
            c_nom, r_ = peor.index[0], peor.iloc[0]
            recs.append((f"⚙️ Revisión sugerida: {c_nom}",
                         f"ASR de {fmt_asr(r_['ASR'])} con {fmt_num(r_['CALLS'])} llamadas. "
                         f"Evaluar causa raíz y posible ajuste de ruteo o penalidad contractual."))
    if not ent.empty:
        malas_ent = (ent.groupby("RUTA").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                     .query("ASR < 30 and CALLS > 100"))
        if not malas_ent.empty:
            recs.append(("⚙️ Rutas con ASR bajo en Entrante",
                         f"{len(malas_ent)} rutas entrantes con ASR &lt; 30%. Evaluar restricciones "
                         f"de CPS o condicionar a mejoras de calidad."))
    if not recs:
        item("✅ Sin acciones urgentes", "No se detectaron anomalías relevantes para recomendar acción hoy.")
    for t, txt in recs:
        item(t, txt)

    box_section("📈 Métricas de Referencia Rápida")
    metricas = pd.DataFrame({
        "MÉTRICA": ["Total Intentos", "Llamadas Exitosas", "Total Minutos", "ASR Promedio",
                    "Rutas Activas", "Rutas ASR > 80%", "Rutas ASR < 30%"],
        "SALIENTE": [
            fmt_num(calls_sal), fmt_num(ok_sal), fmt_num(min_sal, 0), fmt_asr(asr_sal),
            str(sal["RUTA"].nunique()),
            str((sal.groupby("RUTA")["ASR"].mean() >= 80).sum()) if not sal.empty else "0",
            str((sal.groupby("RUTA")["ASR"].mean() < 30).sum()) if not sal.empty else "0",
        ],
        "ENTRANTE": [
            fmt_num(calls_ent), fmt_num(ok_ent), fmt_num(min_ent, 0), fmt_asr(asr_ent),
            str(ent["RUTA"].nunique()),
            str((ent.groupby("RUTA")["ASR"].mean() >= 80).sum()) if not ent.empty else "0",
            str((ent.groupby("RUTA")["ASR"].mean() < 30).sum()) if not ent.empty else "0",
        ],
    })
    _diario_tabla(metricas, DIARIO_C["navy_title"],
                  [("MÉTRICA", "MÉTRICA", None), ("SALIENTE", "SALIENTE", None),
                   ("ENTRANTE", "ENTRANTE", None)],
                  "tbl_diario_metricas", estado_col=False)


def build_diario_xlsx(dia, fecha_str):
    """
    Genera el reporte Diario completo en Excel (bytes), replicando el diseno del
    'Darwin Daily Report' original: hojas Resumen, Entrante, Saliente, POIS, Insights,
    mismos colores, encabezados y semaforo de ESTADO por ASR.
    """
    WHITE, DARK = "FFFFFFFF", "FF1A1A2E"
    THIN = Side(style="thin", color="FFB0B0B0")
    BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    def estado_fill(v):
        if v >= 80: return "FFD9F2D9"
        if v >= 50: return "FFFFF2CC"
        return "FFFFD7D7"

    def estado_txt(v):
        if v >= 80: return "✅ BUENO"
        if v >= 50: return "⚠️ REGULAR"
        return "🔴 BAJO"

    def title_row(ws, ncols, texto, bg, size=16):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=texto)
        c.font = Font(name="Arial", size=size, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 21

    def subtitle_row(ws, ncols, texto, bg):
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=texto)
        c.font = Font(name="Arial", size=10, color="FFAAAAAA")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center")

    ent = dia[dia["TIPO"] == "ENTRANTE"]
    sal = dia[dia["TIPO"] == "SALIENTE"]

    wb = Workbook()
    wb.remove(wb.active)

    # ── Resumen ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")
    title_row(ws, 10, f"DARWIN · DAILY TRAFFIC REPORT · {fecha_str}", "FF1F3864")
    subtitle_row(ws, 10, "Tráfico Saliente · Tráfico Entrante · Análisis de Rutas", "FF0D1B2A")

    for i, h in enumerate(["", "LLAMADAS TOTALES", "LLAMADAS OK", "MINUTOS", "ASR"]):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=9, bold=True)
        c.alignment = Alignment(horizontal="center")

    def resumen_fila(row, label, bg, data):
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Arial", size=8, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right")
        tot = data["CALLS_TOTAL"].sum()
        vals = [tot, data["CALLS_OK"].sum(), data["MINUTOS"].sum(),
                (data["CALLS_OK"].sum() / tot) if tot else 0]
        for j, (v, f) in enumerate(zip(vals, ["#,##0", "#,##0", "#,##0.00", "0.00%"])):
            cc = ws.cell(row=row, column=2 + j, value=v)
            cc.font = Font(name="Arial", size=9, bold=True, color=DARK)
            cc.number_format = f
            cc.alignment = Alignment(horizontal="center")

    resumen_fila(5, "ENTRANTES", "FF375623", ent)
    resumen_fila(6, "SALIENTES", "FFC55A11", sal)

    top_sal = (sal.groupby("CARRIER").agg(CALLS_TOTAL=("CALLS_TOTAL", "sum"),
               MINUTOS=("MINUTOS", "sum"), ASR=("ASR", "mean")).nlargest(10, "MINUTOS").reset_index())
    top_ent = (ent.groupby("CARRIER").agg(CALLS_TOTAL=("CALLS_TOTAL", "sum"),
               MINUTOS=("MINUTOS", "sum"), ASR=("ASR", "mean")).nlargest(10, "MINUTOS").reset_index())

    ws.merge_cells("B8:E8"); c = ws["B8"]; c.value = "🔴  TOP 10 CARRIERS · SALIENTE (por Minutos)"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor="FFC55A11")
    ws.merge_cells("G8:J8"); c = ws["G8"]; c.value = "🟢  TOP 10 CARRIERS · ENTRANTE (por Minutos)"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor="FF375623")

    for start_col in (2, 7):
        for i, h in enumerate(["CARRIER", "CALLS TOTAL", "MINUTOS", "ASR %"]):
            c = ws.cell(row=9, column=start_col + i, value=h)
            c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor="FF2F5597")
            c.alignment = Alignment(horizontal="center")

    for tbl, start_col in [(top_sal, 2), (top_ent, 7)]:
        for i, (_, r) in enumerate(tbl.iterrows()):
            row = 10 + i
            vals = [r["CARRIER"], r["CALLS_TOTAL"], r["MINUTOS"], (r["ASR"] or 0) / 100]
            fmts = ["General", "#,##0", "#,##0.00", "0.00%"]
            for j, (v, f) in enumerate(zip(vals, fmts)):
                c = ws.cell(row=row, column=start_col + j, value=v)
                c.font = Font(name="Arial", size=9, color=DARK)
                c.fill = PatternFill("solid", fgColor="FFEBF3FB")
                c.number_format = f
                c.alignment = Alignment(horizontal="left" if j == 0 else "right")

    for col, w in {"A": 10, "B": 34, "C": 16, "D": 14, "E": 10, "F": 3,
                   "G": 34, "H": 16, "I": 14, "J": 10}.items():
        ws.column_dimensions[col].width = w

    # ── Entrante / Saliente / POIS ──────────────────────────────────────────
    def hoja_trafico(nombre, data, titulo_tipo, bg, bg_sub, origen_dest):
        ws2 = wb.create_sheet(nombre)
        title_row(ws2, 8, f"TRÁFICO {titulo_tipo} · DARWIN · {fecha_str}", bg)
        subtitle_row(ws2, 8,
            f"Total rutas: {data['RUTA'].nunique():,} | "
            f"Total llamadas: {int(data['CALLS_TOTAL'].sum()):,} | "
            f"Total minutos: {int(data['MINUTOS'].sum()):,}", bg_sub)
        cols = [f"CARRIER {origen_dest}", f"RUTA {origen_dest}", "ASR %",
                "CALLS TOTAL", "CALLS OK", "MINUTOS", "ACD (min)", "ESTADO"]
        for i, h in enumerate(cols):
            c = ws2.cell(row=3, column=i + 1, value=h)
            c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center")

        tabla = (data.groupby(["CARRIER", "RUTA"])
                 .agg(ASR=("ASR", "mean"), CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                      CALLS_OK=("CALLS_OK", "sum"), MINUTOS=("MINUTOS", "sum"),
                      AVG_CALL=("AVG_CALL", "mean"))
                 .reset_index().sort_values("CALLS_TOTAL", ascending=False))
        for i, (_, r) in enumerate(tabla.iterrows()):
            row = 4 + i
            asr_v = r["ASR"] or 0
            vals = [r["CARRIER"], r["RUTA"], asr_v / 100, r["CALLS_TOTAL"], r["CALLS_OK"],
                    r["MINUTOS"], r["AVG_CALL"], estado_txt(asr_v)]
            fmts = ["General", "General", "0.00%", "#,##0", "#,##0", "#,##0.00", "0.00", "General"]
            for j, (v, f) in enumerate(zip(vals, fmts)):
                c = ws2.cell(row=row, column=1 + j, value=v)
                c.font = Font(name="Arial", size=9, color=DARK)
                c.number_format = f
                if j == 7:
                    c.fill = PatternFill("solid", fgColor=estado_fill(asr_v))
                    c.alignment = Alignment(horizontal="center")
        for col, w in zip("ABCDEFGH", [36, 22, 10, 14, 12, 14, 12, 14]):
            ws2.column_dimensions[col].width = w
        return tabla

    hoja_trafico("Entrante", ent, "ENTRANTE", "FF375623", "FF1A3A0F", "ORIGEN")
    hoja_trafico("Saliente", sal, "SALIENTE", "FFC55A11", "FF7B2D00", "DESTINO")
    poi = dia[dia["RUTA"].str.contains("POI", case=False, na=False) & (dia["TIPO"] == "SALIENTE")]
    if not poi.empty:
        hoja_trafico("POIS", poi, "POIS", "FFC55A11", "FF7B2D00", "DESTINO")

    # ── Insights ─────────────────────────────────────────────────────────────
    wsi = wb.create_sheet("Insights")
    title_row(wsi, 10, f"INSIGHTS & ALERTAS · DARWIN · {fecha_str}", "FF0D1B2A")
    subtitle_row(wsi, 10, "Análisis automático del tráfico diario · Rutas críticas · "
                          "Recomendaciones operativas", "FF0D1B2A")
    r = [4]

    def sec(texto):
        wsi.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=10)
        c = wsi.cell(row=r[0], column=1, value=texto)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor="FF2F5597")
        r[0] += 1

    def linea(titulo, texto, bg="FFFFFFFF"):
        wsi.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=2)
        c1 = wsi.cell(row=r[0], column=1, value=titulo)
        c1.font = Font(name="Arial", size=9, bold=True, color=DARK)
        c1.fill = PatternFill("solid", fgColor=bg)
        wsi.merge_cells(start_row=r[0], start_column=3, end_row=r[0], end_column=10)
        c2 = wsi.cell(row=r[0], column=3, value=texto)
        c2.font = Font(name="Arial", size=9, color=DARK)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
        r[0] += 1

    calls_sal, ok_sal, min_sal = sal["CALLS_TOTAL"].sum(), sal["CALLS_OK"].sum(), sal["MINUTOS"].sum()
    calls_ent, ok_ent, min_ent = ent["CALLS_TOTAL"].sum(), ent["CALLS_OK"].sum(), ent["MINUTOS"].sum()
    asr_sal = sal["ASR"].mean() if not sal.empty else 0
    asr_ent = ent["ASR"].mean() if not ent.empty else 0

    sec("📋  RESUMEN EJECUTIVO DEL DÍA")
    linea("📞  Volumen total de tráfico",
          f"Se procesaron {int(calls_sal):,} intentos de llamadas salientes, completando "
          f"{int(ok_sal):,} llamadas exitosas ({int(min_sal):,} minutos). ASR promedio "
          f"saliente {asr_sal:.1f}% y entrante {asr_ent:.1f}%.")
    linea("📡  Cobertura de rutas",
          f"Se registraron {sal['RUTA'].nunique()} rutas salientes activas y "
          f"{ent['RUTA'].nunique()} rutas entrantes.")

    sec("🚨  ALERTAS CRÍTICAS · ASR BAJO")
    for tipo, data in [("Saliente", sal), ("Entrante", ent)]:
        if data.empty: continue
        malas = (data.groupby("RUTA").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                 .query("ASR < 30 and CALLS > 100").sort_values("CALLS", ascending=False))
        if not malas.empty:
            lista = " | ".join(f"{ru} ({a:.0f}%)" for ru, a in
                                zip(malas.index[:5], malas["ASR"].iloc[:5]))
            linea(f"⚡  {tipo}: {len(malas)} rutas con ASR < 30%", lista, "FFFFD7D7")

    sec("🏆  RUTAS DESTACADAS · MEJOR PERFORMANCE")
    for tipo, data in [("SAL", sal), ("ENT", ent)]:
        if data.empty: continue
        buenas = (data.groupby(["RUTA", "CARRIER"])
                  .agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"), MIN=("MINUTOS", "sum"))
                  .query("ASR >= 90 and CALLS > 500")
                  .sort_values("CALLS", ascending=False).head(3))
        for (ruta, carrier), row_ in buenas.iterrows():
            linea(f"✅  {tipo} · {ruta}",
                  f"Carrier: {carrier} | ASR: {row_['ASR']:.1f}% | {int(row_['CALLS']):,} llamadas | "
                  f"{int(row_['MIN']):,} min — Ruta estable.", "FFD9F2D9")

    sec("📊  ANÁLISIS POR CARRIER PRINCIPAL")
    if not sal.empty:
        top_c = (sal.groupby("CARRIER").agg(CALLS=("CALLS_TOTAL", "sum"), MIN=("MINUTOS", "sum"),
                 ASR=("ASR", "mean")).nlargest(3, "MIN"))
        for carrier, row_ in top_c.iterrows():
            linea(f"📶  {carrier} · Saliente",
                  f"{int(row_['CALLS']):,} llamadas | {int(row_['MIN']):,} min | "
                  f"ASR promedio: {row_['ASR']:.1f}%")

    sec("💡  RECOMENDACIONES OPERATIVAS")
    n_recs = r[0]
    if not sal.empty:
        peor = (sal.groupby("CARRIER").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                .query("CALLS > 1000").nsmallest(1, "ASR"))
        if not peor.empty:
            c_nom, r_ = peor.index[0], peor.iloc[0]
            linea(f"⚙️  Revisión sugerida: {c_nom}",
                  f"ASR de {r_['ASR']:.1f}% con {int(r_['CALLS']):,} llamadas. "
                  f"Evaluar causa raíz y ruteo.")
    if not ent.empty:
        malas_ent = (ent.groupby("RUTA").agg(ASR=("ASR", "mean"), CALLS=("CALLS_TOTAL", "sum"))
                     .query("ASR < 30 and CALLS > 100"))
        if not malas_ent.empty:
            linea("⚙️  Rutas con ASR bajo en Entrante",
                  f"{len(malas_ent)} rutas entrantes con ASR < 30%. Evaluar restricciones "
                  f"de CPS o condicionar a mejoras de calidad.")
    if r[0] == n_recs:
        linea("✅  Sin acciones urgentes", "No se detectaron anomalías relevantes hoy.")

    sec("📈  MÉTRICAS DE REFERENCIA RÁPIDA")
    for i, h in enumerate(["MÉTRICA", "SALIENTE", "ENTRANTE", "OBSERVACIÓN"]):
        col = i + 1 if i < 3 else 4
        c = wsi.cell(row=r[0], column=col, value=h)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor="FF2F5597")
    r[0] += 1
    metricas = [
        ("Total Intentos", int(calls_sal), int(calls_ent), "Volumen total del día"),
        ("Llamadas Exitosas", int(ok_sal), int(ok_ent), "Completadas efectivamente"),
        ("Total Minutos", int(min_sal), int(min_ent), "Minutos cursados"),
        ("ASR Promedio", f"{asr_sal:.1f}%", f"{asr_ent:.1f}%", "Tasa de respuesta promedio"),
        ("Rutas Activas", sal["RUTA"].nunique(), ent["RUTA"].nunique(), "Rutas con tráfico"),
    ]
    for m in metricas:
        for j, v in enumerate(m):
            c = wsi.cell(row=r[0], column=j + 1, value=v)
            c.font = Font(size=9, color=DARK)
        r[0] += 1

    for col, w in zip("ABCDEFGHIJ", [24, 14, 30, 10, 10, 10, 10, 10, 10, 10]):
        wsi.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def page_diario(df, filtro):
    """
    Hoja 'Diario': replica el reporte Excel 'Darwin Daily Report' dentro del dashboard,
    con 5 sub-hojas (Resumen, Clientes, Proveedores, POIs, Insights) usando los mismos
    colores y formato del Excel original, generado con la informacion del dia cargado.
    Boton de descarga del reporte diario completo al final de la sub-hoja Resumen.
    """
    # Dia de referencia: respeta el filtro "Dia" del sidebar si esta en un solo dia;
    # si no, usa el ultimo dia cargado en los archivos (reporte diario mas reciente).
    if filtro["FECHA"].nunique() == 1:
        fecha_dia = filtro["FECHA"].iloc[0]
    else:
        fecha_dia = df["FECHA"].max()
    dia = df[df["FECHA"] == fecha_dia].copy()
    fecha_str = pd.to_datetime(fecha_dia).strftime("%d/%m/%Y")

    ent = dia[dia["TIPO"] == "ENTRANTE"]
    sal = dia[dia["TIPO"] == "SALIENTE"]

    sub = st.tabs(["📋 Resumen", "📞 Clientes", "📡 Proveedores", "📍 POIs", "💡 Insights"])

    # ══════════════════ RESUMEN ══════════════════
    with sub[0]:
        _diario_title(f"DARWIN · DAILY TRAFFIC REPORT · {fecha_str}",
                      "Tráfico Saliente · Tráfico Entrante · Análisis de Rutas",
                      DIARIO_C["navy_title"], DIARIO_C["navy_sub"])

        tot_ent = ent["CALLS_TOTAL"].sum()
        tot_sal = sal["CALLS_TOTAL"].sum()
        asr_ent_dia = (ent["CALLS_OK"].sum() / tot_ent * 100) if tot_ent else float("nan")
        asr_sal_dia = (sal["CALLS_OK"].sum() / tot_sal * 100) if tot_sal else float("nan")

        st.markdown(f"""
        <table class="ht" style="width:100%;margin-bottom:16px;">
        <thead><tr>
            <th style="background:{DIARIO_C['blue_header']} !important;color:#fff;"></th>
            <th style="background:{DIARIO_C['blue_header']} !important;color:#fff;">LLAMADAS TOTALES</th>
            <th style="background:{DIARIO_C['blue_header']} !important;color:#fff;">LLAMADAS OK</th>
            <th style="background:{DIARIO_C['blue_header']} !important;color:#fff;">MINUTOS</th>
            <th style="background:{DIARIO_C['blue_header']} !important;color:#fff;">ASR</th>
        </tr></thead>
        <tbody>
        <tr>
            <td style="background:{DIARIO_C['green_title']};color:#fff;font-weight:800;text-align:right;">ENTRANTES</td>
            <td style="background:{DIARIO_C['row_alt']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(tot_ent)}</td>
            <td style="background:{DIARIO_C['row_alt']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(ent['CALLS_OK'].sum())}</td>
            <td style="background:{DIARIO_C['row_alt']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(ent['MINUTOS'].sum(),2)}</td>
            <td style="background:{DIARIO_C['row_alt']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_asr(asr_ent_dia)}</td>
        </tr>
        <tr>
            <td style="background:{DIARIO_C['orange_title']};color:#fff;font-weight:800;text-align:right;">SALIENTES</td>
            <td style="background:{DIARIO_C['row_base']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(tot_sal)}</td>
            <td style="background:{DIARIO_C['row_base']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(sal['CALLS_OK'].sum())}</td>
            <td style="background:{DIARIO_C['row_base']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_num(sal['MINUTOS'].sum(),2)}</td>
            <td style="background:{DIARIO_C['row_base']};text-align:center;color:{DIARIO_C['row_txt']};font-weight:700;">{fmt_asr(asr_sal_dia)}</td>
        </tr>
        </tbody></table>
        """, unsafe_allow_html=True)

        top_sal = (sal.groupby("CARRIER").agg(CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                   MINUTOS=("MINUTOS", "sum"), ASR=("ASR", "mean"))
                   .nlargest(10, "MINUTOS").reset_index()) if not sal.empty else pd.DataFrame()
        top_ent = (ent.groupby("CARRIER").agg(CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                   MINUTOS=("MINUTOS", "sum"), ASR=("ASR", "mean"))
                   .nlargest(10, "MINUTOS").reset_index()) if not ent.empty else pd.DataFrame()

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f'<div style="background:{DIARIO_C["orange_title"]};color:#fff;'
                        f'font-weight:800;padding:6px 12px;border-radius:6px 6px 0 0;'
                        f'font-size:0.82rem;">🔴 TOP 10 CARRIERS · SALIENTE (por Minutos)</div>',
                        unsafe_allow_html=True)
            if not top_sal.empty:
                _diario_tabla(top_sal, DIARIO_C["blue_header"],
                              [("CARRIER", "CARRIER", None), ("CALLS_TOTAL", "CALLS TOTAL", fmt_num),
                               ("MINUTOS", "MINUTOS", lambda x: fmt_num(x, 2)),
                               ("ASR", "ASR %", fmt_asr)],
                              "tbl_diario_top_sal", estado_col=False)
            else:
                st.info("Sin datos salientes para el día.")
        with col2:
            st.markdown(f'<div style="background:{DIARIO_C["green_title"]};color:#fff;'
                        f'font-weight:800;padding:6px 12px;border-radius:6px 6px 0 0;'
                        f'font-size:0.82rem;">🟢 TOP 10 CARRIERS · ENTRANTE (por Minutos)</div>',
                        unsafe_allow_html=True)
            if not top_ent.empty:
                _diario_tabla(top_ent, DIARIO_C["blue_header"],
                              [("CARRIER", "CARRIER", None), ("CALLS_TOTAL", "CALLS TOTAL", fmt_num),
                               ("MINUTOS", "MINUTOS", lambda x: fmt_num(x, 2)),
                               ("ASR", "ASR %", fmt_asr)],
                              "tbl_diario_top_ent", estado_col=False)
            else:
                st.info("Sin datos entrantes para el día.")

        st.markdown("<br>", unsafe_allow_html=True)
        excel_bytes = build_diario_xlsx(dia, fecha_str)
        st.download_button(
            "⬇️  Descargar Reporte Diario Completo (Excel)",
            data=excel_bytes,
            file_name=f"Darwin_Daily_Report_{pd.to_datetime(fecha_dia).strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ══════════════════ CLIENTES (Entrante) ══════════════════
    with sub[1]:
        if ent.empty:
            st.info("Sin datos de tráfico entrante para el día seleccionado.")
        else:
            _diario_title(
                f"TRÁFICO ENTRANTE · DARWIN · {fecha_str}",
                f"Total rutas: {ent['RUTA'].nunique()} | "
                f"Total llamadas: {fmt_num(ent['CALLS_TOTAL'].sum())} | "
                f"Total minutos: {fmt_num(ent['MINUTOS'].sum())}",
                DIARIO_C["green_title"], DIARIO_C["green_sub"])
            tabla_ent = (ent.groupby(["CARRIER", "RUTA"])
                         .agg(ASR=("ASR", "mean"), CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                              CALLS_OK=("CALLS_OK", "sum"), MINUTOS=("MINUTOS", "sum"),
                              AVG_CALL=("AVG_CALL", "mean"))
                         .reset_index().sort_values("CALLS_TOTAL", ascending=False))
            _diario_tabla(tabla_ent, DIARIO_C["green_title"],
                          [("CARRIER", "CARRIER ORIGEN", None), ("RUTA", "RUTA ORIGEN", None),
                           ("ASR", "ASR %", fmt_asr), ("CALLS_TOTAL", "CALLS TOTAL", fmt_num),
                           ("CALLS_OK", "CALLS OK", fmt_num),
                           ("MINUTOS", "MINUTOS", lambda x: fmt_num(x, 2)),
                           ("AVG_CALL", "ACD (min)", lambda x: fmt_num(x, 2))],
                          "tbl_diario_clientes")

    # ══════════════════ PROVEEDORES (Saliente) ══════════════════
    with sub[2]:
        if sal.empty:
            st.info("Sin datos de tráfico saliente para el día seleccionado.")
        else:
            _diario_title(
                f"TRÁFICO SALIENTE · DARWIN · {fecha_str}",
                f"Total rutas: {sal['RUTA'].nunique()} | "
                f"Total llamadas: {fmt_num(sal['CALLS_TOTAL'].sum())} | "
                f"Total minutos: {fmt_num(sal['MINUTOS'].sum())}",
                DIARIO_C["orange_title"], DIARIO_C["orange_sub"])
            tabla_sal = (sal.groupby(["CARRIER", "RUTA"])
                         .agg(ASR=("ASR", "mean"), CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                              CALLS_OK=("CALLS_OK", "sum"), MINUTOS=("MINUTOS", "sum"),
                              AVG_CALL=("AVG_CALL", "mean"))
                         .reset_index().sort_values("CALLS_TOTAL", ascending=False))
            _diario_tabla(tabla_sal, DIARIO_C["orange_title"],
                          [("CARRIER", "CARRIER DESTINO", None), ("RUTA", "RUTA DESTINO", None),
                           ("ASR", "ASR %", fmt_asr), ("CALLS_TOTAL", "CALLS TOTAL", fmt_num),
                           ("CALLS_OK", "CALLS OK", fmt_num),
                           ("MINUTOS", "MINUTOS", lambda x: fmt_num(x, 2)),
                           ("AVG_CALL", "ACD (min)", lambda x: fmt_num(x, 2))],
                          "tbl_diario_proveedores")

    # ══════════════════ POIs ══════════════════
    with sub[3]:
        poi = dia[dia["RUTA"].str.contains("POI", case=False, na=False) & (dia["TIPO"] == "SALIENTE")]
        if poi.empty:
            st.info("No se encontraron rutas POI para el día seleccionado.")
        else:
            _diario_title(f"TRÁFICO POIS · DARWIN · {fecha_str}",
                          f"Total rutas: {poi['RUTA'].nunique()} | "
                          f"Total llamadas: {fmt_num(poi['CALLS_TOTAL'].sum())}",
                          DIARIO_C["orange_title"], DIARIO_C["orange_sub"])
            tabla_poi = (poi.groupby(["CARRIER", "RUTA"])
                         .agg(ASR=("ASR", "mean"), CALLS_TOTAL=("CALLS_TOTAL", "sum"),
                              CALLS_OK=("CALLS_OK", "sum"), MINUTOS=("MINUTOS", "sum"),
                              AVG_CALL=("AVG_CALL", "mean"))
                         .reset_index().sort_values("CALLS_TOTAL", ascending=False))
            _diario_tabla(tabla_poi, DIARIO_C["orange_title"],
                          [("CARRIER", "CARRIER DESTINO", None), ("RUTA", "RUTA DESTINO", None),
                           ("ASR", "ASR %", fmt_asr), ("CALLS_TOTAL", "CALLS TOTAL", fmt_num),
                           ("CALLS_OK", "CALLS OK", fmt_num),
                           ("MINUTOS", "MINUTOS", lambda x: fmt_num(x, 2)),
                           ("AVG_CALL", "ACD (min)", lambda x: fmt_num(x, 2))],
                          "tbl_diario_poi")
            tot_poi = poi["CALLS_TOTAL"].sum()
            asr_tot = (poi["CALLS_OK"].sum() / tot_poi * 100) if tot_poi else float("nan")
            st.markdown(f"""
            <div class="insight-box" style="margin-top:10px;">
                <b>TOTAL POIs</b> — Calls: <b>{fmt_num(poi['CALLS_TOTAL'].sum())}</b> ·
                Calls OK: <b>{fmt_num(poi['CALLS_OK'].sum())}</b> ·
                Minutos: <b>{fmt_num(poi['MINUTOS'].sum(),2)}</b> ·
                ASR: <b>{fmt_asr(asr_tot)}</b>
            </div>""", unsafe_allow_html=True)

    # ══════════════════ INSIGHTS ══════════════════
    with sub[4]:
        _diario_title(f"INSIGHTS & ALERTAS · DARWIN · {fecha_str}",
                      "Análisis automático del tráfico diario · Rutas críticas · "
                      "Recomendaciones operativas",
                      DIARIO_C["navy_title"], DIARIO_C["navy_sub"])
        _diario_insights(sal, ent)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="IPLAN | ASR Dashboard",
        page_icon="📡",layout="wide",
        initial_sidebar_state="expanded",  # siempre abierta al inicio
    )
    inject_css()

    st.markdown(f"""
    <div class="iplan-header">
        <img src="{LOGO_URL}" alt="IPLAN">
        <div class="iplan-header-text">
            <h1>ASR Dashboard</h1>
            <span>TRAFICO ENT &amp; SAL · RUTAS · CLIENTES · POIs · PREFIX CC</span>
        </div>
        <div class="header-badge">NOC WHOLESALE</div>
    </div>
    """, unsafe_allow_html=True)

    # Forzar sidebar expandida siempre
    st.markdown(
        """<style>
        [data-testid="stSidebar"]{display:block!important;visibility:visible!important;
        transform:translateX(0)!important;opacity:1!important;}
        </style>""",
        unsafe_allow_html=True,
    )

    with st.spinner("Cargando datos..."):
        df = load_all_data()

    if df.empty:
        st.error(f"No se encontraron datos.\n• {CARPETA_ENTRANTE}\n• {CARPETA_SALIENTE}")
        st.stop()

    df["FECHA"] = pd.to_datetime(df["FECHA"])
    df["MES"]   = df["FECHA"].dt.to_period("M").astype(str)

    with st.sidebar:
        st.markdown(f'<div style="text-align:center;padding:4px 0 10px;">'
                    f'<img src="{LOGO_URL}" style="height:22px;filter:brightness(0) invert(1);"></div>',
                    unsafe_allow_html=True)
        st.markdown("---")
        tipo_opts   = st.multiselect("Trafico",["ENTRANTE","SALIENTE"],
                                     default=["ENTRANTE","SALIENTE"])
        meses_disp  = sorted(df["MES"].unique().tolist(),reverse=True)
        mes_sel     = st.multiselect("Mes",meses_disp,
                                     default=meses_disp[:1] if meses_disp else [])
        df_mes      = df[df["MES"].isin(mes_sel)] if mes_sel else df
        fechas_d    = sorted(df_mes["FECHA"].dt.date.unique().tolist(),reverse=True)
        fecha_sel   = st.selectbox("Dia",["Todos"]+[str(f) for f in fechas_d])
        carrier_sel = st.multiselect("Carrier",
                                     sorted(df_mes["CARRIER"].dropna().unique().tolist()),
                                     default=[])
        ruta_sel    = st.multiselect("Ruta",
                                     sorted(df_mes["RUTA"].dropna().unique().tolist()),
                                     default=[])
        st.markdown("---")
        if st.button("Refrescar",use_container_width=True):
            st.cache_data.clear(); st.rerun()
        st.markdown(f"""
        <div class="stats-box">
            <b>Dias:</b> {df['FECHA'].nunique()}<br>
            <b>Registros:</b> {len(df):,}<br>
            <b>Ultimo:</b> {df['FECHA'].max().strftime('%d/%m/%Y')}
        </div>""", unsafe_allow_html=True)

    filtro = df.copy()
    if tipo_opts:   filtro = filtro[filtro["TIPO"].isin(tipo_opts)]
    if mes_sel:     filtro = filtro[filtro["MES"].isin(mes_sel)]
    if fecha_sel != "Todos":
        filtro = filtro[filtro["FECHA"].dt.date == pd.to_datetime(fecha_sel).date()]
    if carrier_sel: filtro = filtro[filtro["CARRIER"].isin(carrier_sel)]
    if ruta_sel:    filtro = filtro[filtro["RUTA"].isin(ruta_sel)]

    if filtro.empty:
        st.warning("Sin datos para los filtros seleccionados."); st.stop()

    tabs = st.tabs(["Resumen","Clientes","Proveedores","POIs IPLAN","Prefix CC","Diario"])
    with tabs[0]: page_resumen(df,filtro)
    with tabs[1]: page_clientes(df,filtro)
    with tabs[2]: page_salientes(df,filtro)
    with tabs[3]: page_poi(df,filtro)
    with tabs[4]: page_prefix_cc(df,filtro)
    with tabs[5]: page_diario(df,filtro)

if __name__ == "__main__":
    main()
